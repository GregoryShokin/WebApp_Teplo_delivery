"""Первичная загрузка реестра ОС по итогам фотоинвентаризации 2026.

Источник — «Реестр_ОС_и_амортизация_2026-07-01.xlsx», лист «Реестр ОС»: 110 позиций,
149 единиц, 2 313 150 ₽, единая дата ввода 01.07.2026. Реестр собран по фотофиксации
(склад 09.06.2026, ТК Черникова 27.07.2026, дополнительная съёмка 30.07.2026) и подтверждён
владельцем построчно.

ОДНА КАРТОЧКА = ОДНА ФИЗИЧЕСКАЯ ЕДИНИЦА. Позиции с количеством больше одного разворачиваются
(24 позиции дают 39 добавочных карточек). Иначе инвентарный номер на группу из трёх ларей
бессмыслен, а списать один из трёх нечем.

ДАТА ВВОДА — 01.07.2026, а не дата описи. Решение владельца 30.07.2026: всё это оборудование
существовало и РАБОТАЛО на 1 июля, ничего с тех пор не покупалось. Инвентаризация лишь
задокументировала уже эксплуатируемое имущество, поэтому её дата к вводу в эксплуатацию
отношения не имеет. Дата берётся из колонки реестра, здесь не хардкодится.

Порог признания ОС (10 000 ₽) к этой загрузке НЕ применяется — решение владельца 2026-07-30:
имущество уже оценено и ставится на баланс целиком. Порог работает против будущих покупок.

СПИ карточкам НЕ проставляется: срок берётся из категории (``resolve_useful_life``). В реестре
срок у всех позиций одного типа одинаков, поэтому дублировать его в карточке незачем — а если
владелец потом поменяет срок категории, график пересчитается для всей категории разом.

Скрипт идемпотентен по ``source_ref`` («Черникова №27», «Склад №12 (2 из 3)»): повторный
прогон ничего не задваивает и честно скажет, сколько строк уже было.

    docker exec -w /app/apps/api <api-контейнер> \
        python -m app.scripts.import_fixed_assets_registry --file /app/<реестр>.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import asyncio
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import AsyncSessionLocal
from app.models import AssetCategory, FixedAsset, Location
from app.services.fixed_assets import create_asset

SHEET = "Реестр ОС"

# Колонки листа (0-based) — порядок задан генератором реестра и в файле не меняется.
COL_NUMBER = 0
COL_SOURCE = 1
COL_INVENTORY_REF = 2
COL_NAME = 3
COL_BRAND = 4
COL_QTY = 5
COL_UNIT_COST = 6
COL_TYPE = 8
COL_COMMISSIONED = 10

# «Не работающее оборудование» в реестре стоит в колонке типа рядом с настоящими категориями,
# но категорией НЕ является: в модели это статус карточки. Все три единицы — холодильные
# витрины (POLAIR CN105S, четвёртая настольная витрина, демонтированный элемент витрины),
# поэтому категорию ставим настоящую, а неамортизируемость обеспечивает статус.
NOT_WORKING_TYPE = "Не работающее оборудование"
NOT_WORKING_CATEGORY = "Холодильное/морозильное оборудование"

# Колонка «Источник» реестра → помещение в реестре помещений.
LOCATION_BY_SOURCE = {"Черникова": "Черникова", "Склад": "Склад"}

# Источники, где оборудование лежит в РЕЗЕРВЕ, а не работает. Решение владельца 2026-07-30:
# «Ни одно из основных средств на складе не работает, всё стоит в резерве». Амортизация при
# этом идёт: исправная техника в запасе стареет и устаревает так же, как в работе.
RESERVE_SOURCES = frozenset({"Склад"})


class RegistryImportError(RuntimeError):
    """Загрузку нельзя продолжать: реестр не сходится с состоянием базы."""


@dataclass(frozen=True)
class Unit:
    """Одна физическая единица — будущая карточка ОС."""

    name: str
    brand_model: str | None
    unit_cost: Decimal
    type_name: str
    source: str
    source_ref: str
    commissioned_on: date
    status: str


def _as_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise RegistryImportError(f"Не дата в колонке «Дата ввода»: {value!r}")


def _clean(value: object) -> str | None:
    text = str(value).strip() if value is not None else ""
    return text or None


def read_units(path: Path) -> list[Unit]:
    """Развернуть строки реестра в отдельные единицы.

    Строка «ИТОГО» в конце листа не импортируется: у неё пустой номер позиции.
    """
    workbook = load_workbook(path, data_only=True, read_only=True)
    if SHEET not in workbook.sheetnames:
        raise RegistryImportError(f"В книге нет листа «{SHEET}»: {workbook.sheetnames}")
    sheet = workbook[SHEET]

    units: list[Unit] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[COL_NUMBER] is None:
            continue
        source = str(row[COL_SOURCE]).strip()
        inventory_ref = row[COL_INVENTORY_REF]
        name = str(row[COL_NAME]).strip()
        quantity = int(row[COL_QTY])
        unit_cost = Decimal(str(row[COL_UNIT_COST]))
        type_name = str(row[COL_TYPE]).strip()
        commissioned_on = _as_date(row[COL_COMMISSIONED])
        base_ref = f"{source} №{inventory_ref}"
        # Статус — это «работает ли», а не «где стоит». Реестр инвентаризации их не различал,
        # поэтому берём из источника: на складе НИЧЕГО не работает, всё лежит в резерве
        # (подтверждено владельцем 2026-07-30). Поставить всем «в работе» значило бы объявить
        # работающим полмиллиона рублей оборудования, которое выручку не производит.
        # Неработающее остаётся неработающим где угодно: сломанное — не резерв.
        if type_name == NOT_WORKING_TYPE:
            status = "not_working"
        else:
            status = "in_storage" if source in RESERVE_SOURCES else "in_use"

        for index in range(1, quantity + 1):
            # Номер экземпляра внутри позиции: три ларя из одной строки описи — три разные
            # карточки, и по ссылке должно быть видно, какая из них какая.
            ref = base_ref if quantity == 1 else f"{base_ref} ({index} из {quantity})"
            units.append(
                Unit(
                    name=name,
                    brand_model=_clean(row[COL_BRAND]),
                    unit_cost=unit_cost,
                    type_name=type_name,
                    source=source,
                    source_ref=ref,
                    commissioned_on=commissioned_on,
                    status=status,
                )
            )
    workbook.close()
    return units


async def _categories(session: AsyncSession, units: list[Unit]) -> dict[str, AssetCategory]:
    rows = (await session.scalars(select(AssetCategory))).all()
    by_name = {row.name: row for row in rows}

    required = {
        NOT_WORKING_CATEGORY if unit.type_name == NOT_WORKING_TYPE else unit.type_name
        for unit in units
    }
    missing = sorted(required - by_name.keys())
    if missing:
        raise RegistryImportError(
            "В справочнике категорий ОС нет: "
            + ", ".join(f"«{name}»" for name in missing)
            + ". Прогони миграции (0221 сеет категории) и повтори."
        )
    return by_name


async def _locations(session: AsyncSession, units: list[Unit]) -> dict[str, Location]:
    rows = (await session.scalars(select(Location))).all()
    by_name = {row.name: row for row in rows}

    resolved: dict[str, Location] = {}
    unknown: set[str] = set()
    for source in {unit.source for unit in units}:
        target = LOCATION_BY_SOURCE.get(source)
        location = by_name.get(target) if target else None
        if location is None:
            unknown.add(source)
        else:
            resolved[source] = location
    if unknown:
        raise RegistryImportError(
            "В реестре помещений нет: "
            + ", ".join(f"«{name}»" for name in sorted(unknown))
            + ". Заведи помещение и повтори — иначе имущество не сопоставить с расходами."
        )
    return resolved


async def run(session: AsyncSession, path: Path, *, dry_run: bool) -> dict[str, object]:
    units = read_units(path)
    categories = await _categories(session, units)
    locations = await _locations(session, units)

    known_refs = set(
        (
            await session.scalars(
                select(FixedAsset.source_ref).where(FixedAsset.source_ref.is_not(None))
            )
        ).all()
    )

    created = 0
    skipped = 0
    total = Decimal("0.00")
    by_category: Counter[str] = Counter()

    for unit in units:
        total += unit.unit_cost
        if unit.source_ref in known_refs:
            skipped += 1
            continue
        category_name = (
            NOT_WORKING_CATEGORY if unit.type_name == NOT_WORKING_TYPE else unit.type_name
        )
        category = categories[category_name]
        location = locations[unit.source]
        if not dry_run:
            await create_asset(
                session,
                name=unit.name,
                initial_cost=unit.unit_cost,
                category_id=category.id,
                # Рыночная оценка на дату инвентаризации, а не сумма платежа: историю покупок
                # для этих объектов не восстанавливаем.
                valuation_basis="market",
                valued_on=unit.commissioned_on,
                commissioned_on=unit.commissioned_on,
                status=unit.status,
                location=unit.source,
                location_id=location.id,
                brand_model=unit.brand_model,
                source_ref=unit.source_ref,
            )
        known_refs.add(unit.source_ref)
        created += 1
        by_category[category_name] += 1

    if not dry_run:
        await session.commit()

    return {
        "units": len(units),
        "created": created,
        "skipped": skipped,
        "total": total,
        "by_category": by_category,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Залить реестр ОС инвентаризации 2026 в базу.")
    parser.add_argument("--file", required=True, type=Path, help="Путь к xlsx с реестром")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Посчитать и показать итог, ничего не записывая",
    )
    return parser.parse_args()


async def async_main() -> None:
    args = parse_args()
    if not args.file.exists():
        raise SystemExit(f"Файл не найден: {args.file}")

    async with AsyncSessionLocal() as session:
        result = await run(session, args.file, dry_run=args.dry_run)

    mode = "ПРОБНЫЙ ПРОГОН (ничего не записано)" if args.dry_run else "ЗАГРУЖЕНО"
    print(f"{mode}")
    print(f"  единиц в реестре : {result['units']}")
    print(f"  создано карточек : {result['created']}")
    print(f"  уже было         : {result['skipped']}")
    print(f"  сумма реестра    : {result['total']:,.2f} ₽".replace(",", " "))
    print("  по категориям:")
    for name, count in sorted(result["by_category"].items(), key=lambda item: -item[1]):
        print(f"    {count:>3}  {name}")


def main() -> None:
    asyncio.run(async_main())


if __name__ == "__main__":
    main()
