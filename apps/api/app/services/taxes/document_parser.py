"""Разбор налоговых документов от налогового агента (бухгалтер на аутсорсе, askad02@mail.ru).

Агент присылает стандартные унифицированные формы — это ключ к автоматизации:

* **платёжные поручения** (форма 0401060, ``.docx``): ЕНП, УСН, допвзнос 1%. Дают сумму,
  КБК, получателя, назначение и срок. Тип платежа берётся из имени файла и заголовка —
  «банк не говорит, что внутри платежа», а агент говорит прямо;
* **ведомости** (форма Т-53, ``.xls``): фактические выплаты по сотрудникам за период;
* **ПД (налог)** (``.xls``, «0,2 %»): взнос на травматизм в СФР.

``.docx`` — это zip+xml, читается стандартной библиотекой без зависимостей. ``.xls`` (старый
бинарный BIFF) требует ``xlrd``; ``.xlsx`` читается через ``openpyxl`` — формат выбирается
по сигнатуре содержимого в :func:`open_workbook`.

ВАЖНО про доверие: сумма и КБК из платёжки достоверны. Тип платежа (``tax_kind``) выводится
из ИМЕНИ ФАЙЛА, которое агент пишет вручную — формат стабильный, но человеческий. Поэтому
разбор возвращает ``needs_review=True`` там, где уверенности нет, а не угадывает молча.
"""

from __future__ import annotations

import contextlib
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

# КБК → назначение. Единый КБК ЕНП покрывает УСН, НДФЛ, взносы «за себя» — их не различить
# по КБК, только по имени файла. КБК травматизма отдельный.
KBK_ENP = "18201061201010000510"
KBK_INJURY = "79710212000061000160"

# Классификация типа платежа по ключевым словам в имени файла / заголовке документа.
# Порядок важен: «1%» проверяется раньше «УСН», потому что оба могут встретиться.
_KIND_PATTERNS: tuple[tuple[str, str], ...] = (
    (r"1\s*%|1%\s*св|допвзнос|1\s*процент", "contrib_extra_1pct"),
    (r"\bусн\b", "usn_advance"),
    (r"0[.,]2\s*%|травматизм|несчастн", "contrib_injury"),
    # «СВ ИП 1 кв.docx» — фиксированные взносы ИП «за себя» (платятся поквартально по ¼).
    (r"св\s+ип|взнос[а-я]*\s+ип", "contrib_fixed"),
    # «ЕНП» и «ЕНС» — одно и то же в письмах бухгалтера: смешанный НДФЛ+взносы, разнос
    # даёт оборотка. Паттерн «ЕНС» добавлен 27.07.2026 после сверки с выпиской: платёжка
    # «ЕНС до 27.03» (19 460,93) — реальный зарплатный ЕНП за февраль, а классификатор
    # относил её к допвзносу 1% и исказил бы вычет при продвижении.
    (r"\bен[пс]\b", "enp_payroll"),
)


@dataclass(frozen=True)
class PaymentOrderDoc:
    """Разобранное платёжное поручение (форма 0401060)."""

    amount: Decimal | None
    kbk: str | None
    recipient: str | None  # 'fns' | 'sfr'
    tax_kind: str | None  # ключ TaxPayment.kind или 'enp_payroll' (смешанный) / None
    due_date: date | None
    purpose: str | None
    period_hint: str | None  # 'q1'|'h1'|'9m'|'year' | 'YYYY-MM' — из имени файла, если понятно
    needs_review: bool
    review_reasons: list[str] = field(default_factory=list)
    source_filename: str | None = None
    # Откуда взят период: 'document' — из самого документа, 'letter' — подставлен по месяцу
    # письма (каскад), None — не определялся. Владелец должен видеть, что месяц выведен,
    # а не прочитан: подставленный месяц — допущение, и в спорном случае его правят руками.
    period_source: str | None = None


@dataclass(frozen=True)
class PayrollRow:
    tab_number: str | None
    employee: str
    amount: Decimal


@dataclass(frozen=True)
class PayrollStatementDoc:
    """Разобранная платёжная ведомость (форма Т-53)."""

    doc_number: str | None
    payout_kind: str | None  # 'advance' | 'salary' — из имени файла
    period_start: date | None
    period_end: date | None
    rows: list[PayrollRow]
    total: Decimal
    needs_review: bool
    review_reasons: list[str] = field(default_factory=list)
    source_filename: str | None = None


# ── книги Excel: .xls (xlrd) и .xlsx (openpyxl) под одним интерфейсом ────────


class _XlsxSheet:
    """Лист .xlsx под интерфейс листа xlrd (name/nrows/ncols/cell_value)."""

    def __init__(self, name: str, rows: list[tuple]) -> None:
        self.name = name
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(r) for r in rows), default=0)

    def cell_value(self, row: int, col: int) -> object:
        cells = self._rows[row]
        value = cells[col] if col < len(cells) else None
        if value is None:  # xlrd отдаёт пустую строку, openpyxl — None
            return ""
        if isinstance(value, datetime):  # даты — текстом, как пишут в наших документах
            return value.strftime("%d.%m.%Y")
        return value


class _XlsxWorkbook:
    """Книга .xlsx под интерфейс xlrd (sheets/sheet_names/sheet_by_name)."""

    def __init__(self, data: bytes) -> None:
        from openpyxl import load_workbook  # локальный импорт: тяжёлый только для Excel

        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        try:
            self._sheets = [
                _XlsxSheet(ws.title, [tuple(row) for row in ws.iter_rows(values_only=True)])
                for ws in wb.worksheets
            ]
        finally:
            wb.close()

    def sheets(self) -> list[_XlsxSheet]:
        return list(self._sheets)

    def sheet_names(self) -> list[str]:
        return [s.name for s in self._sheets]

    def sheet_by_name(self, name: str) -> _XlsxSheet:
        return next(s for s in self._sheets if s.name == name)


def open_workbook(data: bytes) -> object:
    """Книга Excel по СИГНАТУРЕ содержимого (не по расширению — имена рукописные):
    zip (PK) → .xlsx через openpyxl, иначе → .xls через xlrd."""
    if data[:4] == b"PK\x03\x04":
        return _XlsxWorkbook(data)
    import xlrd  # локальный импорт: тяжёлый только для .xls-веток

    return xlrd.open_workbook(file_contents=data)


# ── docx ─────────────────────────────────────────────────────────────────────


def _docx_text(data: bytes) -> str:
    """Текст docx по абзацам. Стандартная библиотека, без зависимостей."""
    with zipfile.ZipFile(BytesIO(data)) as z:
        xml = z.read("word/document.xml").decode("utf-8", "replace")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"<w:tab[^>]*/>", " ", xml)
    return re.sub(r"<[^>]+>", "", xml)


# ── pdf ──────────────────────────────────────────────────────────────────────
#
# Бухгалтер шлёт те же формы то в Word/Excel, то печатью в PDF (одна и та же ведомость
# ВЕД-14 пришла 22.07 в .xls и 27.07 в .pdf). Формат выбирается по СИГНАТУРЕ содержимого,
# как в :func:`open_workbook`, — расширения в письмах рукописные и не всегда честные.

PDF_SIGNATURE = b"%PDF"


def is_pdf(data: bytes) -> bool:
    return data[:4] == PDF_SIGNATURE


def _pdf_text(data: bytes) -> str:
    """Текст цифрового PDF (``pypdf``). Скан вернёт пусто — разбор честно уйдёт в проверку."""
    from io import BytesIO as _BytesIO  # локально: pypdf тяжёлый и нужен только pdf-ветке

    from pypdf import PdfReader

    reader = PdfReader(_BytesIO(data))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


# Фамилия и инициалы в PDF приходят отдельными словами («ВОДОЛАЗОВА В.С.»), а Excel отдаёт их
# одной ячейкой. Склеиваем обратно — иначе строка сотрудника не опознаётся (см. _NAME_RE).
_PDF_SURNAME_RE = re.compile(r"^[А-ЯЁ][А-ЯЁа-яё]+(?:-[А-ЯЁ][А-ЯЁа-яё]+)?$")
_PDF_INITIALS_RE = re.compile(r"^[А-ЯЁ]\.\s*[А-ЯЁ]\.?$")


def _pdf_grid(data: bytes) -> list[list[str]]:
    """PDF → «таблица» строк, как её видят Т-53-хелперы: строка текста → список ячеек."""
    grid: list[list[str]] = []
    for line in _pdf_text(data).split("\n"):
        cells = [token for token in line.split() if token]
        if not cells:
            continue
        merged: list[str] = []
        for cell in cells:
            if (
                merged
                and _PDF_INITIALS_RE.match(cell)
                and _PDF_SURNAME_RE.match(merged[-1])
            ):
                merged[-1] = f"{merged[-1]} {cell}"
                continue
            merged.append(cell)
        grid.append(merged)
    return grid


def _parse_amount(text: str) -> Decimal | None:
    """Сумма из формата «105628-00» (рубли-копейки). Берём в блоке «Сумма»/итога.

    Ноль (0-00) — валидная сумма (нулевая платёжка-заглушка), поэтому отличаем «не нашли»
    (None) от «ноль» (Decimal 0).
    """
    m = re.search(r"\b(\d{1,9})-(\d{2})\b", text)
    if not m:
        return None
    return Decimal(m.group(1)) + Decimal(m.group(2)) / 100


def _parse_kbk(text: str) -> str | None:
    for m in re.finditer(r"\b(\d{20})\b", text):
        code = m.group(1)
        if code.startswith(("182", "797")):
            return code
    return None


def _parse_due_date(text: str, filename: str, default_year: int) -> date | None:
    """Срок «до 28.07» / «до 25.09» из имени файла или заголовка. Год — из полной даты
    в тексте, иначе из ``default_year`` (год письма)."""
    for source in (filename, text):
        m = re.search(r"до\s+(\d{1,2})[.\s](\d{1,2})(?:[.\s](\d{2,4}))?", source, re.I)
        if m:
            day, month = int(m.group(1)), int(m.group(2))
            year = _normalize_year(m.group(3), default_year)
            try:
                return date(year, month, day)
            except ValueError:
                continue
    return None


def _normalize_year(raw: str | None, default_year: int) -> int:
    if not raw:
        return default_year
    y = int(raw)
    return 2000 + y if y < 100 else y


def _classify_kind(filename: str, header: str) -> str | None:
    haystack = f"{filename} {header}".lower()
    for pattern, kind in _KIND_PATTERNS:
        if re.search(pattern, haystack):
            return kind
    return None


def _period_hint(filename: str, header: str) -> str | None:
    text = f"{filename} {header}".lower()
    if re.search(r"1\s*кв|1\s*квартал", text):
        return "q1"
    if re.search(r"2\s*кв|полугод|2\s*квартал", text):
        return "h1"
    if re.search(r"3\s*кв|9\s*мес", text):
        return "9m"
    if re.search(r"год|4\s*кв", text):
        return "year"
    m = re.search(r"\b(0?[1-9]|1[0-2])[.\-](\d{4})\b", text)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    return None


def parse_payment_order(
    data: bytes, *, filename: str = "", default_year: int | None = None
) -> PaymentOrderDoc:
    """Разобрать платёжное поручение (docx или печать в pdf, форма 0401060)."""
    text = _pdf_text(data) if is_pdf(data) else _docx_text(data)
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    header = lines[0] if lines else ""
    year = default_year or _guess_year(text) or date.today().year

    amount = _parse_amount(text)
    kbk = _parse_kbk(text)
    recipient = _recipient_from(text, kbk)
    tax_kind = _classify_kind(filename, header)
    due = _parse_due_date(text, filename, year)
    purpose = _purpose_from(lines)
    period = _period_hint(filename, header)

    reasons: list[str] = []
    if amount is None:
        reasons.append("не распознана сумма")
    if kbk is None:
        reasons.append("не распознан КБК")
    if tax_kind is None:
        reasons.append("не распознан вид платежа по имени файла/заголовку")
    if due is None:
        reasons.append("не распознан срок уплаты")
    # ЕНП (смешанный НДФЛ+взносы) НЕ уходит в needs_review: разнос делается из оборотки
    # бухгалтера (rebuild_payroll_enp_split), а не из уведомления. Платёжка тут — справочная
    # (подтверждает уплаченную единой суммой ЕНП), отдельного ручного подтверждения не требует.

    return PaymentOrderDoc(
        amount=amount,
        kbk=kbk,
        recipient=recipient,
        tax_kind=tax_kind,
        due_date=due,
        purpose=purpose,
        period_hint=period,
        needs_review=bool(reasons),
        review_reasons=reasons,
        source_filename=filename or None,
    )


_INJURY_PERIOD_RE = re.compile(r"МС\.(0?[1-9]|1[0-2])\.(20\d{2})", re.I)


def _injury_period_month(cells: list[str]) -> tuple[int, int] | None:
    """Месяц начисления из поля периода платёжки «МС.07.2026» → (год, месяц)."""
    for cell in cells:
        m = _INJURY_PERIOD_RE.search(cell)
        if m:
            return int(m.group(2)), int(m.group(1))
    return None


def _month_before(year: int, month: int) -> tuple[int, int]:
    """Предыдущий месяц — обратная операция к :func:`_injury_due`."""
    return (year - 1, 12) if month == 1 else (year, month - 1)


def _injury_due(year: int, month: int) -> date:
    """Срок уплаты взноса на травматизм за месяц N — 15 число N+1 (125-ФЗ, ст. 22)."""
    return date(year + 1, 1, 15) if month == 12 else date(year, month + 1, 15)


def parse_injury_payment(
    data: bytes,
    *,
    filename: str = "",
    default_year: int | None = None,
    received_on: date | None = None,
) -> PaymentOrderDoc:
    """Разобрать платёжку взноса на травматизм — «Форма ПД (налог)» (.xls) в СФР.

    Это единая платёжка «0,2 %» по всем сотрудникам (не ведомость Т-53 и не форма 0401060):
    внутри — сумма, КБК травматизма (797…), получатель ОСФР по региону. Раньше такой .xls
    молча уходил в Т-53-парсер, где «не находил сотрудников» и вис в «нужна проверка».

    ``received_on`` — дата письма; последняя ступень каскада определения месяца (см. ниже).
    """
    wb = open_workbook(data)
    cells: list[str] = []
    for sheet in wb.sheets():
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                value = str(sheet.cell_value(r, c)).strip()
                if value:
                    cells.append(value)
    text = " ".join(cells)
    year = default_year or _guess_year(text) or date.today().year

    amount = _injury_amount(cells)
    kbk = next(
        (c for c in cells if re.fullmatch(r"\d{20}", c) and c.startswith("797")),
        None,
    )

    reasons: list[str] = []
    if amount is None:
        reasons.append("не распознана сумма взноса")
    if kbk is None:
        reasons.append("не распознан КБК травматизма")

    # Срока уплаты в форме ПД нет — его задаёт закон: до 15 числа месяца, следующего за
    # месяцем начисления (125-ФЗ, ст. 22). Без месяца обязательство получало «срок» = дата
    # письма и на следующий день краснело просрочкой.
    #
    # КАСКАД определения месяца — три ступени, сверху вниз:
    #   1. поле периода «МС.MM.YYYY» в самом документе — единственный достоверный источник;
    #   2. «до DD.MM» в имени файла или тексте — это СРОК, а не месяц начисления: по 125-ФЗ
    #      срок стоит на месяц позже начисления, поэтому месяц берём минус один. Раньше срок
    #      клали прямо в период, и _injury_due прибавлял ещё месяц — обязательство уезжало на
    #      месяц вперёд («до 15.09» давало период 2026-09 и срок 15.10, то есть пени);
    #   3. МЕСЯЦ ПИСЬМА. С апреля 2026 форма ПД приходит без поля периода (ячейка пуста), а
    #      файлы разных месяцев побайтно неразличимы по ячейкам — читать больше нечего.
    #      По всей истории, где поле было (дек-2025 … мар-2026, 5 из 5), месяц начисления
    #      совпадал с месяцем письма: агент выписывает извещение в том же месяце, 17-26 числа.
    #      Это ДОПУЩЕНИЕ, а не факт из документа, поэтому помечаем ``period_source='letter'``
    #      — владелец видит бейдж и может поправить месяц руками.
    #
    # Зачем каскад вообще: без месяца распознанное вырождается в пару (сумма, КБК) —
    # одинаковую у всех извещений на 100 ₽, — и контентный дедуп приёмника (_duplicate_intake_id)
    # гасит свежий документ как повтор прошлого. Так 19.08.2026 потерялся август.
    month = _injury_period_month(cells)
    period_year, period_month = month if month else (None, None)
    period_source = "document" if period_month else None
    if period_month is None:
        pp_date = _parse_due_date(text, filename, year)
        # Год в «до 15.01» не пишут, и _parse_due_date подставляет год ПИСЬМА: у декабрьского
        # письма «до 15.01» становилось 15.01 того же года, месяц начисления уезжал в прошлый
        # декабрь, а обязательство — в чужой налоговый год. Срок не может быть раньше письма.
        if pp_date is not None and received_on is not None and pp_date < received_on:
            pp_date = pp_date.replace(year=pp_date.year + 1)
        if pp_date is not None:
            period_year, period_month = _month_before(pp_date.year, pp_date.month)
            period_source = "filename"
    if period_month is None and received_on is not None:
        period_year, period_month = received_on.year, received_on.month
        period_source = "letter"
    if period_month is None:
        reasons.append("не распознан месяц начисления")
    # Без месяца срока нет: единственный его источник — _parse_due_date — на этой ветке
    # уже вернул None, иначе месяц бы нашёлся ступенью 2.
    due = _injury_due(period_year, period_month) if period_year and period_month else None

    return PaymentOrderDoc(
        amount=amount,
        kbk=kbk,
        recipient="sfr",
        tax_kind="contrib_injury",
        due_date=due,
        purpose="Страховые взносы на травматизм (в СФР)",
        # Каскад исчерпан — период ПУСТОЙ, а не «какой-нибудь». _period_hint здесь
        # возвращал 'year' из подстроки «ГОД» в «ВОЛГОДОНСК» (адрес плательщика) — и
        # документ снова получал вырожденное распознанное, совпадающее с прошлыми
        # извещениями на ту же сумму. Пустой период честнее: документ уйдёт владельцу.
        period_hint=(
            f"{period_year}-{period_month:02d}" if period_year and period_month else None
        ),
        needs_review=bool(reasons),
        review_reasons=reasons,
        source_filename=filename or None,
        period_source=period_source,
    )


def _injury_amount(cells: list[str]) -> Decimal | None:
    """Сумма взноса — ячейка сразу после метки «Сумма» (формат 57.14 / 100,00 / 100)."""

    def _num(raw: str) -> Decimal | None:
        s = raw.replace(" ", "").replace(",", ".")
        if re.fullmatch(r"\d{1,9}(\.\d+)?", s):
            with contextlib.suppress(ArithmeticError):
                return Decimal(s)
        return None

    for i, cell in enumerate(cells):
        if "сумма" in cell.lower():
            for nxt in cells[i + 1 : i + 3]:
                value = _num(nxt)
                if value is not None and value > 0:
                    return value
    return None


def _recipient_from(text: str, kbk: str | None) -> str | None:
    if kbk == KBK_INJURY or "ОСФР" in text or "СФР" in text.upper():
        return "sfr"
    if kbk == KBK_ENP or "Казначейство" in text or "ФНС" in text:
        return "fns"
    return None


def _purpose_from(lines: list[str]) -> str | None:
    for i, line in enumerate(lines):
        if line.startswith("Назначение платежа") and i > 0:
            return lines[i - 1]
    for line in lines:
        if line in ("ЕНП.", "ЕНП"):
            return "Единый налоговый платеж"
    return None


def _guess_year(text: str) -> int | None:
    m = re.search(r"\b(20\d{2})\b", text)
    return int(m.group(1)) if m else None


# ── xls (ведомость Т-53) ─────────────────────────────────────────────────────

# Строка сотрудника в Т-53: «ФАМИЛИЯ И.О.» + сумма. Фамилия в ведомостях идёт ЗАГЛАВНЫМИ,
# инициалы — обязательно с точками (это отличает сотрудника от подписи руководителя,
# где инициалы без точек).
_NAME_RE = re.compile(
    r"^[А-ЯЁ][А-ЯЁа-яё]+(?:-[А-ЯЁ][А-ЯЁа-яё]+)?\s+[А-ЯЁ]\.\s*[А-ЯЁ]\.?$"
)


def parse_payroll_statement(data: bytes, *, filename: str = "") -> PayrollStatementDoc:
    """Разобрать платёжную ведомость (xls/xlsx или печать в pdf, форма Т-53)."""
    if is_pdf(data):
        grid = _pdf_grid(data)
    else:
        wb = open_workbook(data)
        grid = []
        for sheet in wb.sheets():
            for r in range(sheet.nrows):
                grid.append([str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)])

    doc_number = _find_doc_number(grid)
    period_start, period_end = _find_period(grid)
    rows = _find_employee_rows(grid)
    total = sum((row.amount for row in rows), Decimal("0"))
    payout_kind = _payout_kind(filename)

    reasons: list[str] = []
    if not rows:
        reasons.append("не найдено ни одной строки сотрудника")
    if period_start is None:
        reasons.append("не распознан расчётный период")
    if payout_kind is None:
        reasons.append("не распознан тип выплаты (аванс/зарплата) по имени файла")

    return PayrollStatementDoc(
        doc_number=doc_number,
        payout_kind=payout_kind,
        period_start=period_start,
        period_end=period_end,
        rows=rows,
        total=total,
        needs_review=bool(reasons),
        review_reasons=reasons,
        source_filename=filename or None,
    )


def _payout_kind(filename: str) -> str | None:
    low = filename.lower()
    if "аванс" in low:
        return "advance"
    if "зп" in low or "зарплат" in low:
        return "salary"
    return None


def _find_doc_number(grid: list[list[str]]) -> str | None:
    """Номер ведомости формата «20-13». Первое совпадение — это и есть номер документа."""
    for row in grid:
        for cell in row:
            if re.fullmatch(r"\d{1,3}-\d{1,3}", cell):
                return cell
    return None


def _find_period(grid: list[list[str]]) -> tuple[date | None, date | None]:
    dates: list[date] = []
    for row in grid:
        for cell in row:
            m = re.fullmatch(r"(\d{2})\.(\d{2})\.(\d{4})", cell)
            if m:
                with contextlib.suppress(ValueError):
                    dates.append(date(int(m.group(3)), int(m.group(2)), int(m.group(1))))
    if not dates:
        return None, None
    # Расчётный период — минимальная и максимальная даты в документе (1-е и последнее число).
    first_of_month = [d for d in dates if d.day == 1]
    start = min(first_of_month) if first_of_month else min(dates)
    end = max(dates)
    return start, end


def _find_employee_rows(grid: list[list[str]]) -> list[PayrollRow]:
    rows: list[PayrollRow] = []
    seen: set[tuple[str, str]] = set()
    for row in grid:
        name = next((c for c in row if _NAME_RE.match(c)), None)
        if not name:
            continue
        amount = _row_amount(row)
        if amount is None:
            continue
        tab = _row_tab_number(row)
        key = (name, str(amount))
        if key in seen:
            continue
        seen.add(key)
        rows.append(PayrollRow(tab_number=tab, employee=name, amount=amount))
    return rows


def _row_amount(row: list[str]) -> Decimal | None:
    """Денежная сумма в строке: формат 20986.00 (xls хранит числа как float-строки)."""
    for cell in row:
        m = re.fullmatch(r"(\d{3,7})\.(\d{2})", cell)
        if m:
            return Decimal(m.group(1)) + Decimal(m.group(2)) / 100
    return None


def _row_tab_number(row: list[str]) -> str | None:
    for cell in row:
        m = re.fullmatch(r"(\d{2,4})(?:\.0)?", cell)
        if m and 100 <= int(m.group(1)) <= 9999:
            return m.group(1)
    return None


# ── xls (сальдо-оборотная ведомость по зарплате, лист 'л1') ───────────────────

# Оборотка — не платёжная ведомость Т-53, а РАСЧЁТНАЯ: помесячная раскладка начислений и
# удержаний по каждому сотруднику (оклад → НДФЛ → аванс → взносы → травматизм → вычет → к
# выплате). Это эталон для разноса зарплатного ЕНП (НДФЛ vs взносы за работников) и для
# монитора зарплатного критерия льготы по НДС. Значимый лист называется 'л1'.
#
# Колонки формы стабильны ПО ПОЗИЦИЯМ, а подписи «плавают» (в июле поз. «аванс» подписана
# «УДЕРЖАНИЯ ПО КАССЕ»), поэтому читаем по ИНДЕКСУ колонки, а не по тексту заголовка. Все
# значения в файле хранятся ТЕКСТОМ («50000.00», «13595.93»), поэтому парсим Decimal из строки.
_TURNOVER_SHEET = "л1"
# Индексы колонок листа 'л1' (проверено на файлах мая/июня/июля 2026).
_TCOL_TAB = 1  # табельный номер
_TCOL_NAME = 2  # ФИО
_TCOL_OKLAD = 4  # оклад (тариф)
_TCOL_DAYS = 5  # отработано дней
_TCOL_ACCRUED = 6  # начислено всего
_TCOL_NDFL = 8  # НДФЛ («налог на доходы»)
_TCOL_ADVANCE = 9  # аванс / «удержания по кассе»
_TCOL_CONTRIB = 10  # страховые взносы за работника (СФР)
_TCOL_INJURY = 11  # травматизм 0,2 %
_TCOL_DEDUCTION = 12  # налоговые вычеты (напр. на ребёнка)
_TCOL_TOPAY = 13  # к выплате

_MONTHS_RU: dict[str, int] = {
    "ЯНВАРЬ": 1, "ФЕВРАЛЬ": 2, "МАРТ": 3, "АПРЕЛЬ": 4, "МАЙ": 5, "ИЮНЬ": 6,
    "ИЮЛЬ": 7, "АВГУСТ": 8, "СЕНТЯБРЬ": 9, "ОКТЯБРЬ": 10, "НОЯБРЬ": 11, "ДЕКАБРЬ": 12,
}


@dataclass(frozen=True)
class TurnoverRow:
    """Строка сотрудника оборотки — суммы за расчётный месяц. None ≠ 0 («ячейка пуста»)."""

    tab_number: str | None
    employee: str
    oklad: Decimal | None
    days: Decimal | None
    accrued: Decimal | None  # начислено
    ndfl: Decimal | None
    advance: Decimal | None  # уже выплаченный аванс / удержания по кассе
    contributions: Decimal | None  # страховые взносы за работника (СФР)
    injury: Decimal | None  # травматизм
    deduction: Decimal | None  # налоговые вычеты
    to_pay: Decimal | None  # к выплате


@dataclass(frozen=True)
class TurnoverStatementDoc:
    """Разобранная сальдо-оборотная ведомость по зарплате за один месяц."""

    year: int | None
    month: int | None
    period_code: str | None  # 'YYYY-MM'
    rows: list[TurnoverRow]
    accrued_total: Decimal
    ndfl_total: Decimal
    contributions_total: Decimal
    injury_total: Decimal
    needs_review: bool
    review_reasons: list[str] = field(default_factory=list)
    source_filename: str | None = None


def _turnover_decimal(value: object) -> Decimal | None:
    """Число из ячейки оборотки. Значения — текст «50000.00»/«13595,93», иногда пусто/float."""
    text = repr(value) if isinstance(value, (int, float)) else str(value)
    text = text.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except (ArithmeticError, ValueError):
        return None


def _turnover_value(sheet: object, row: int, col: int) -> Decimal | None:
    if col >= sheet.ncols:  # type: ignore[attr-defined]
        return None
    return _turnover_decimal(sheet.cell_value(row, col))  # type: ignore[attr-defined]


def _turnover_period(grid: list[list[str]]) -> tuple[int | None, int | None]:
    """Месяц и год из заголовка «Сальдо-оборотная ведомость по з/плате ИЮЛЬ 2026»."""
    header = " ".join(cell for line in grid[:3] for cell in line).upper()
    month = next((m for name, m in _MONTHS_RU.items() if name in header), None)
    ym = re.search(r"\b(20\d{2})\b", header)
    year = int(ym.group(1)) if ym else None
    return year, month


# ── СВОД по начислениям-удержаниям ───────────────────────────────────────────
# Вторая форма того же содержания: помесячные ИТОГИ без разбивки по сотрудникам.
# Бухгалтер присылает её, когда оборотка не выгружается («за февраль не могу пока выгрузить
# оборотку, в своде тоже есть цифры начисления страховых взносов», письмо 27.07.2026).
_SUMMARY_HEADER_RE = re.compile(r"по\s+начислени\w*[\s\-–]*удержани", re.I)
_SUMMARY_PERIOD_RE = re.compile(r"за\s+(0?[1-9]|1[0-2])\.(20\d{2})")
_SUMMARY_AMOUNT_COL = 2  # «Вид Н-У | Дата | Сумма» — сумма в третьей колонке
_SUMMARY_TOTALS_COL = 3  # блок «Итоги» — сумма в четвёртой
_SUMMARY_ROW_LABEL = "СВОД (все сотрудники)"


def _summary_num(grid: list[list[str]], row: int, col: int) -> Decimal | None:
    if row >= len(grid) or col >= len(grid[row]):
        return None
    return _turnover_decimal(grid[row][col])


def _parse_payroll_summary(
    grid: list[list[str]], *, filename: str
) -> TurnoverStatementDoc:
    """Свод «по начислениям-удержаниям за MM.YYYY» → та же структура, что и оборотка.

    Строк сотрудников в своде нет — суммы уже итоговые, поэтому отдаём ОДНУ строку
    ``СВОД (все сотрудники)`` без табельного номера. Виды опознаём по названию:
    «СТР.СФР» — взносы за работников, «Стр.Тр.» — травматизм, «НалогФЛ» — НДФЛ (строк
    может быть несколько: НДФЛ месяца платится «окнами», в своде он разложен по срокам).
    """
    header = " ".join(cell for line in grid[:4] for cell in line)
    period = _SUMMARY_PERIOD_RE.search(header)
    year = int(period.group(2)) if period else None
    month = int(period.group(1)) if period else None

    accrued = ndfl = contributions = injury = advance = to_pay = oklad = days = None
    for r, line in enumerate(grid):
        label = (line[0] if line else "").strip()
        if not label:
            continue
        low = label.casefold()
        value = _summary_num(grid, r, _SUMMARY_AMOUNT_COL)
        if "оклад" in low and value is not None:
            oklad = (oklad or Decimal("0")) + value
            days = days if days is not None else _summary_num(grid, r, 3)
        elif "стр.сфр" in low and value is not None:
            contributions = (contributions or Decimal("0")) + value
        elif "стр.тр" in low and value is not None:
            injury = (injury or Decimal("0")) + value
        elif "налогфл" in low and value is not None:
            ndfl = (ndfl or Decimal("0")) + value
        elif "зп 1/2м" in low and value is not None:
            advance = (advance or Decimal("0")) + value
        elif "всего начислено" in low:
            accrued = _summary_num(grid, r, _SUMMARY_TOTALS_COL) or accrued
        elif "всего к выдаче" in low:
            to_pay = _summary_num(grid, r, _SUMMARY_TOTALS_COL)
    if accrued is None:
        accrued = oklad

    row = TurnoverRow(
        tab_number=None,
        employee=_SUMMARY_ROW_LABEL,
        oklad=oklad,
        days=days,
        accrued=accrued,
        ndfl=ndfl,
        advance=advance,
        contributions=contributions,
        injury=injury,
        deduction=None,
        to_pay=to_pay,
    )
    reasons: list[str] = []
    if year is None or month is None:
        reasons.append("не распознан месяц/год из заголовка свода")
    if contributions is None:
        reasons.append("в своде не найдены страховые взносы (СТР.СФР)")
    if accrued is None:
        reasons.append("в своде не найдено начислено")
    return TurnoverStatementDoc(
        year=year,
        month=month,
        period_code=f"{year}-{month:02d}" if (year and month) else None,
        rows=[row],
        accrued_total=accrued or Decimal("0"),
        ndfl_total=ndfl or Decimal("0"),
        contributions_total=contributions or Decimal("0"),
        injury_total=injury or Decimal("0"),
        needs_review=bool(reasons),
        review_reasons=reasons,
        source_filename=filename or None,
    )


def parse_turnover_statement(data: bytes, *, filename: str = "") -> TurnoverStatementDoc:
    """Разобрать сальдо-оборотную ведомость по зарплате (xls/xlsx, лист 'л1').

    Если листа 'л1' нет, но документ похож на СВОД по начислениям-удержаниям — читаем его:
    у свода те же итоги месяца, только без разбивки по сотрудникам.
    """
    wb = open_workbook(data)
    if _TURNOVER_SHEET not in wb.sheet_names():
        for name in wb.sheet_names():
            sheet = wb.sheet_by_name(name)
            if not sheet.nrows:
                continue
            grid = [
                [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ]
            head = " ".join(cell for line in grid[:4] for cell in line)
            if _SUMMARY_HEADER_RE.search(head):
                return _parse_payroll_summary(grid, filename=filename)
        return TurnoverStatementDoc(
            year=None, month=None, period_code=None, rows=[],
            accrued_total=Decimal("0"), ndfl_total=Decimal("0"),
            contributions_total=Decimal("0"), injury_total=Decimal("0"),
            needs_review=True,
            review_reasons=[f"не найден лист '{_TURNOVER_SHEET}' — это не оборотка"],
            source_filename=filename or None,
        )

    sheet = wb.sheet_by_name(_TURNOVER_SHEET)
    grid = [
        [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
        for r in range(sheet.nrows)
    ]
    year, month = _turnover_period(grid)
    period_code = f"{year}-{month:02d}" if (year and month) else None

    rows: list[TurnoverRow] = []
    for r in range(sheet.nrows):
        name = grid[r][_TCOL_NAME] if sheet.ncols > _TCOL_NAME else ""
        if not _NAME_RE.match(name):  # строка «Итого» и шапки сюда не попадают
            continue
        rows.append(
            TurnoverRow(
                tab_number=(grid[r][_TCOL_TAB] or None) if sheet.ncols > _TCOL_TAB else None,
                employee=name,
                oklad=_turnover_value(sheet, r, _TCOL_OKLAD),
                days=_turnover_value(sheet, r, _TCOL_DAYS),
                accrued=_turnover_value(sheet, r, _TCOL_ACCRUED),
                ndfl=_turnover_value(sheet, r, _TCOL_NDFL),
                advance=_turnover_value(sheet, r, _TCOL_ADVANCE),
                contributions=_turnover_value(sheet, r, _TCOL_CONTRIB),
                injury=_turnover_value(sheet, r, _TCOL_INJURY),
                deduction=_turnover_value(sheet, r, _TCOL_DEDUCTION),
                to_pay=_turnover_value(sheet, r, _TCOL_TOPAY),
            )
        )

    reasons: list[str] = []
    if not rows:
        reasons.append("не найдено ни одной строки сотрудника")
    if period_code is None:
        reasons.append("не распознан месяц/год из заголовка")
    # Контроль связности: начислено − НДФЛ − аванс = к выплате (если к выплате заполнена).
    for row in rows:
        if row.to_pay is not None and None not in (row.accrued, row.ndfl, row.advance):
            residual = row.accrued - row.ndfl - row.advance
            if abs(residual - row.to_pay) > Decimal("0.01"):
                reasons.append(
                    f"{row.employee}: начислено−НДФЛ−аванс ({residual}) ≠ к выплате ({row.to_pay})"
                )

    def _total(attr: str) -> Decimal:
        return sum((getattr(row, attr) or Decimal("0") for row in rows), Decimal("0"))

    return TurnoverStatementDoc(
        year=year, month=month, period_code=period_code, rows=rows,
        accrued_total=_total("accrued"), ndfl_total=_total("ndfl"),
        contributions_total=_total("contributions"), injury_total=_total("injury"),
        needs_review=bool(reasons), review_reasons=reasons,
        source_filename=filename or None,
    )
