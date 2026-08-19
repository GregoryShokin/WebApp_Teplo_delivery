"""Приём налоговых документов из почты → staging.

IMAP не трогаем: подсовываем фикстуры через инъекцию ``fetch``. Проверяем маршрутизацию
по типу, фильтр отправителя, дедуп и статусы (parsed / needs_review / unsupported).
"""

from __future__ import annotations

import hashlib
import uuid
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.config import get_settings
from app.models.tax import TaxDocumentIntake
from app.services.mail.imap_client import FetchedAttachment, MailAccount
from app.services.taxes.document_ingest import (
    ingest_tax_documents,
    parse_attachment,
    set_intake_review,
)

FIXTURES = Path(__file__).parent / "fixtures" / "taxes"


def _att(
    name: str, filename: str, *, sender: str = "Бухгалтер <askad02@mail.ru>"
) -> FetchedAttachment:
    content = (FIXTURES / name).read_bytes()
    mime = (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        if filename.endswith(".docx")
        else "application/vnd.ms-excel"
    )
    return FetchedAttachment(
        mailbox="corporate",
        message_uid="1",
        message_id=f"<{name}@test>",
        from_addr=sender,
        subject="налоги",
        received_at=datetime(2026, 7, 23, tzinfo=UTC),
        filename=filename,
        mime=mime,
        content=content,
    )


def _fetch_stub(attachments: list[FetchedAttachment]):
    def _fetch(account, *, host, port, lookback_days):  # noqa: ANN001, ARG001
        return list(attachments)

    return _fetch


# Один фиктивный ящик — чтобы приём был герметичным и не зависел от того, настроена ли
# реальная почта в окружении (иначе на два настроенных ящика stub-fetch задвоит вложения).
_TEST_ACCOUNTS = [MailAccount("test", "test@local", "x")]


# ── чистый разбор вложения ───────────────────────────────────────────────────


def test_parse_attachment_routes_payment_order() -> None:
    dtype, status, rec, err = parse_attachment(
        _att("usn_h1_478376.docx", "УСН 2 кв до 28.07.docx")
    )
    assert dtype == "payment_order"
    assert status == "parsed"
    assert rec["amount"] == "478376"
    assert rec["tax_kind"] == "usn_advance"
    assert err is None


def test_parse_attachment_enp_is_parsed_not_review() -> None:
    """ЕНП-платёжка распознаётся (parsed), а не висит в «нужна проверка»: её разнос НДФЛ/взносов
    делает оборотка (rebuild_payroll_enp_split), отдельного ручного подтверждения не требуется."""
    _, status, rec, _ = parse_attachment(_att("enp_payroll_14902.docx", "ЕНП_до 28.07.docx"))
    assert status == "parsed"
    assert rec["tax_kind"] == "enp_payroll"


def test_injury_document_classified_as_payment_order() -> None:
    """Файл травматизма («0,2 %.xls») — это платёжка (payment_order), а не Т-53-ведомость.

    Раньше .xls без «оборот/вед» уходил в Т-53-парсер, где «не находил сотрудников» и виснул
    в «нужна проверка». Теперь ключевые слова травматизма распознаются до .xls-фолбэка.
    """
    from app.services.taxes.document_ingest import _classify_document

    assert _classify_document("0,2 %.xls") == "payment_order"
    assert _classify_document("Травматизм июль.xls") == "payment_order"
    # обычная ведомость по-прежнему Т-53
    assert _classify_document("ВЕД-13 АВАНС 20.07.xls") == "payroll_statement"


def test_injury_amount_extracted_after_label() -> None:
    """Сумма взноса на травматизм берётся из ячейки сразу после метки «Сумма»."""
    from app.services.taxes.document_parser import _injury_amount

    assert _injury_amount(["ИНН", "890307589201", "Сумма", "100.00"]) == Decimal("100.00")
    assert _injury_amount(["Сумма", "57,14"]) == Decimal("57.14")
    assert _injury_amount(["нет метки", "12345"]) is None


def test_parse_attachment_routes_payroll_statement() -> None:
    dtype, status, rec, _ = parse_attachment(
        _att("vedomost_advance_20986.xls", "ВЕД-13 АВАНС 20.07.xls")
    )
    assert dtype == "payroll_statement"
    assert status == "parsed"
    assert rec["total"] == "20986"
    assert rec["rows"][0]["employee"] == "ИВАНОВА И.И."


def test_parse_attachment_routes_turnover_statement() -> None:
    """Оборотка распознаётся отдельным типом (не Т-53) с помесячной раскладкой в recognition."""
    dtype, status, rec, err = parse_attachment(_att("oborotka_07.xls", "ОБОРОТКА 07.xls"))
    assert dtype == "turnover_statement"
    assert status == "parsed"
    assert rec["period_hint"] == "2026-07"
    assert rec["rows"][0]["employee"] == "ИВАНОВА И.И."
    assert rec["rows"][0]["ndfl"] == "6318.00"
    assert rec["rows"][0]["contributions"] == "13595.93"
    assert rec["contributions_total"] == "13595.93"
    assert err is None


def test_parse_attachment_marks_kadr_unsupported() -> None:
    dtype, status, _, _ = parse_attachment(
        _att("vedomost_salary_22696.xls", "Приказ об отпуске.xls")
    )
    assert dtype == "unknown"
    assert status == "unsupported"


def test_kadr_doc_wins_over_extension() -> None:
    """«Приказ … .doc» — кадровый документ, а не платёжка: раньше .doc уводил его в разбор
    docx, где он падал английским «File is not a zip file»."""
    from app.services.taxes.document_ingest import _classify_document

    assert _classify_document("Приказ № 1 12.01.2026 повыш.ЗП.doc") == "unsupported_kadr"
    # Слово «тд» — по границам слова: «отдых» кадровым не считается.
    assert _classify_document("ТД.xls") == "unsupported_kadr"
    assert _classify_document("отдых команды.xls") != "unsupported_kadr"


def _pdf_att(filename: str, *, content: bytes = b"%PDF-1.5\n...") -> FetchedAttachment:
    return FetchedAttachment(
        mailbox="corporate",
        message_uid="2",
        message_id=f"<{filename}@test>",
        from_addr="Бухгалтер <askad02@mail.ru>",
        subject=filename,
        received_at=datetime(2026, 7, 27, tzinfo=UTC),
        filename=filename,
        mime="application/pdf",
        content=content,
    )


def test_pdf_attachment_reaches_tax_contour() -> None:
    """PDF от бухгалтера принимается приёмом налоговых документов.

    27.07.2026 та же ведомость ВЕД-14 пришла в .pdf: фильтр вложений её не пропускал, и
    документ уезжал в контур счетов на оплату — там Т-53 встала в очередь как счёт.
    """
    from email.message import EmailMessage

    from app.services.mail.imap_client import _is_tax_document_part

    part = EmailMessage()
    part.set_content(
        b"%PDF-1.5", maintype="application", subtype="pdf", filename="ВЕД-14 ЗП 05.08.pdf"
    )
    assert _is_tax_document_part(part) is True


def test_parse_attachment_reads_payroll_from_pdf(monkeypatch: pytest.MonkeyPatch) -> None:
    """Ведомость печатью в PDF разбирается так же, как из .xls."""
    from app.services.taxes import document_parser

    monkeypatch.setattr(
        document_parser,
        "_pdf_text",
        lambda data: "1 206 ИВАНОВА И.И. 22696.00\n20-14 05.08.2026 01.07.2026 31.07.2026",
    )

    dtype, status, rec, err = parse_attachment(_pdf_att("ВЕД-14 ЗП 05.08.pdf"))

    assert (dtype, status, err) == ("payroll_statement", "parsed", None)
    assert rec["doc_number"] == "20-14"
    assert rec["total"] == "22696"


def test_turnover_in_pdf_asks_for_excel() -> None:
    """Оборотка читается по индексам колонок листа 'л1' — в PDF сетки нет, просим .xls."""
    dtype, status, rec, err = parse_attachment(_pdf_att("ОБОРОТКА 07.pdf"))

    assert (dtype, status, err) == ("turnover_statement", "unsupported", None)
    assert ".xls" in rec["reason"]


def test_old_doc_format_unsupported_with_russian_reason() -> None:
    """Старый Word (.doc) не читается автоматикой — статус «не поддержан» с русской причиной."""
    dtype, status, rec, err = parse_attachment(
        _att("usn_h1_478376.docx", "Платёжка старая.doc")
    )
    assert dtype == "unknown"
    assert status == "unsupported"
    assert ".docx" in rec["reason"] and "вручную" in rec["reason"]
    assert err is None


def _xlsx_bytes(rows: list[list[object]], sheet_name: str = "Лист1") -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_fixture_as_xlsx(name: str) -> bytes:
    """Существующая xls-фикстура, пересохранённая в .xlsx, — сетка та же."""
    from io import BytesIO

    import xlrd
    from openpyxl import Workbook

    src = xlrd.open_workbook(file_contents=(FIXTURES / name).read_bytes())
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in src.sheets():
        ws = wb.create_sheet(title=sheet.name)
        for r in range(sheet.nrows):
            ws.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_injury_payment_parsed() -> None:
    """Платёжка травматизма в .xlsx читается openpyxl так же, как .xls."""
    from dataclasses import replace

    content = _xlsx_bytes(
        [
            ["Форма ПД (налог)"],
            ["Сумма", "100.00"],
            ["КБК", "79710212000061000160"],
            ["Получатель", "ОСФР по Ростовской области"],
        ]
    )
    att = replace(_att("oborotka_07.xls", "0,2 %.xlsx"), content=content)
    dtype, status, rec, err = parse_attachment(att)
    assert dtype == "payment_order"
    assert status == "parsed"
    assert rec["tax_kind"] == "contrib_injury"
    assert rec["amount"] == "100.00"
    assert rec["kbk"] == "79710212000061000160"
    assert err is None


# ── каскад определения месяца у формы ПД (травматизм) ────────────────────────

# Сетка реального извещения «Форма ПД (налог)», ФИО/ИНН обезличены. Адрес оставлен
# волгодонским намеренно: подстрока «ГОД» внутри «ВОЛГОДОНСК» — та самая ловушка
# _period_hint, из-за которой месячный взнос выглядел годовым платежом.
_PD_GRID: list[list[object]] = [
    ["Рег. № ФСС:6118015592", "Статус", "08", "Форма №  ПД (налог)"],
    ["Извещение", "ФИО", "Адрес"],
    ["ИВАНОВА ИРИНА ИВАНОВНА", "Г. ВОЛГОДОНСК, ЦВЕТОЧНЫЙ Б-Р, 33"],
    ["ИНН", "000000000000", "Сумма", "100.0"],
    ["Получатель", "УФК по Ростовской области (ОСФР по Ростовской области)"],
    ["ИНН", "6163013494", "КПП", "616301001", "Страх.взносы от несчастных случаев"],
    ["79710212000061000160"],
]


def _pd_att(
    *, period_cell: str | None, received: datetime | None, sheet_name: str = "стр.001"
) -> FetchedAttachment:
    """Вложение с извещением ПД: с полем периода «МС.MM.YYYY» или без него.

    ``sheet_name`` разводит БАЙТЫ двух вложений, не трогая ни одной читаемой ячейки — так
    в жизни отличаются файлы разных месяцев (метаданные OLE при одинаковой сетке).
    """
    from dataclasses import replace

    grid = [list(row) for row in _PD_GRID]
    grid.append([period_cell, "0"] if period_cell else ["0"])
    return replace(
        _att("oborotka_07.xls", "0,2 %.xls"),
        content=_xlsx_bytes(grid, sheet_name=sheet_name),
        received_at=received,
    )


def test_injury_period_read_from_document_wins_over_letter() -> None:
    """Ступень 1 каскада: поле «МС.MM.YYYY» в документе сильнее даты письма."""
    att = _pd_att(period_cell="МС.03.2026", received=datetime(2026, 8, 19, tzinfo=UTC))
    _, status, rec, err = parse_attachment(att)

    assert status == "parsed"
    assert rec["period_hint"] == "2026-03"
    assert rec["due_date"] == "2026-04-15"  # 15 число месяца, следующего за начислением
    assert rec["period_source"] == "document"
    assert err is None


def test_injury_period_falls_back_to_letter_month() -> None:
    """Ступень 3: поля периода нет (так с апреля 2026) — месяц берём из даты письма.

    Инцидент 19.08.2026: без этой ступени распознанное вырождалось в пару (сумма, КБК),
    контентный дедуп признавал августовское извещение повтором апрельского и гасил его,
    а взнос за август пропадал из «Налогов» целиком.
    """
    att = _pd_att(period_cell=None, received=datetime(2026, 8, 19, tzinfo=UTC))
    _, status, rec, err = parse_attachment(att)

    assert status == "parsed"
    assert rec["period_hint"] == "2026-08"
    assert rec["due_date"] == "2026-09-15"
    assert rec["period_source"] == "letter"  # месяц выведен — владелец видит бейдж
    assert rec["review_reasons"] == []
    assert err is None


def test_injury_december_letter_rolls_due_date_into_january() -> None:
    """Декабрьское начисление платится до 15 января следующего года (125-ФЗ, ст. 22)."""
    att = _pd_att(period_cell=None, received=datetime(2026, 12, 18, tzinfo=UTC))
    _, _, rec, _ = parse_attachment(att)

    assert rec["period_hint"] == "2026-12"
    assert rec["due_date"] == "2027-01-15"


def test_injury_period_from_filename_deadline_is_month_before() -> None:
    """Ступень 2: «до 15.09» в имени — это СРОК, значит начислено за август, а не за сентябрь.

    Раньше срок клали прямо в период, и _injury_due прибавлял ещё месяц: обязательство
    уезжало на 15.10 — просрочка и пени на ровном месте (125-ФЗ, ст. 22).
    """
    from dataclasses import replace

    att = replace(
        _pd_att(period_cell=None, received=datetime(2026, 8, 19, tzinfo=UTC)),
        filename="0,2 % до 15.09.xls",
    )
    _, status, rec, _ = parse_attachment(att)

    assert status == "parsed"
    assert rec["period_hint"] == "2026-08"
    assert rec["due_date"] == "2026-09-15"
    assert rec["period_source"] == "filename"  # выведено из имени, а не прочитано в документе


def test_injury_without_any_period_source_needs_review() -> None:
    """Каскад исчерпан (нет поля периода и нет даты письма) — документ владельцу, не молча.

    Период при этом ПУСТОЙ, а не «какой-нибудь»: _period_hint возвращал здесь 'year' из
    подстроки «ГОД» в «ВОЛГОДОНСК», и распознанное снова вырождалось до пары (сумма, КБК) —
    то есть совпадало с прошлыми извещениями на ту же сумму.
    """
    att = _pd_att(period_cell=None, received=None)
    _, status, rec, _ = parse_attachment(att)

    assert status == "needs_review"
    assert "не распознан месяц начисления" in rec["review_reasons"]
    assert rec["period_source"] is None
    assert rec["period_hint"] is None, "пустой период честнее выдуманного «года»"


async def test_ingest_does_not_ignore_document_waiting_for_owner(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """Документ в «нужна проверка» не гасится как дубль — у него распознано не всё.

    Два извещения с исчерпанным каскадом дают одинаково бедное распознанное; если гасить
    такие по совпадению, второе исчезнет молча, ни разу не показавшись владельцу.
    """
    first = _pd_att(period_cell=None, received=None)
    second = _pd_att(period_cell=None, received=None, sheet_name="стр.1")
    assert first.sha256 != second.sha256

    async with async_session_factory() as session:
        result = await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([first, second]),
            accounts=_TEST_ACCOUNTS,
        )

    assert result["needs_review"] == 2, result
    assert result.get("ignored", 0) == 0, result


async def test_manual_period_override_clears_derived_mark(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """Владелец поправил месяц — бейдж «месяц по дате письма» больше не про этот документ."""
    att = _pd_att(period_cell=None, received=datetime(2026, 8, 19, tzinfo=UTC))
    async with async_session_factory() as session:
        await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([att]),
            accounts=_TEST_ACCOUNTS,
        )
        intake = (
            await session.execute(select(TaxDocumentIntake))
        ).scalars().one()
        assert intake.recognition["period_source"] == "letter"

        await set_intake_review(
            session,
            intake,
            status="parsed",
            overrides={"period_hint": "2026-07", "due_date": "2026-08-15"},
        )

    assert intake.recognition["period_hint"] == "2026-07"
    assert intake.recognition["period_source"] == "manual"


async def test_ingest_two_months_of_injury_do_not_collapse(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """Два извещения на одну сумму в разные месяцы — два документа, а не «повтор».

    Регрессия 19.08.2026: апрельское и августовское извещения на 100 ₽ давали побайтно
    одинаковое распознанное, и второе уходило в ``ignored`` с ссылкой на первое.
    """
    april = _pd_att(period_cell=None, received=datetime(2026, 4, 17, tzinfo=UTC))
    august = _pd_att(
        period_cell=None, received=datetime(2026, 8, 19, tzinfo=UTC), sheet_name="стр.1"
    )
    assert april.sha256 != august.sha256, "SHA-дедуп не должен схлопнуть их раньше времени"

    async with async_session_factory() as session:
        result = await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([april, august]),
            accounts=_TEST_ACCOUNTS,
        )

    assert result["parsed"] == 2, result
    assert result.get("ignored", 0) == 0, result
    async with async_session_factory() as session:
        rows = (
            await session.execute(
                select(TaxDocumentIntake.status, TaxDocumentIntake.recognition)
            )
        ).all()
    assert sorted(r.recognition["period_hint"] for r in rows) == ["2026-04", "2026-08"]
    assert all("duplicate_of" not in r.recognition for r in rows)


def test_xlsx_turnover_statement_parsed_same_as_xls() -> None:
    """Оборотка в .xlsx даёт ту же раскладку, что и .xls (общий интерфейс книги)."""
    from dataclasses import replace

    att = replace(
        _att("oborotka_07.xls", "ОБОРОТКА 07.xlsx"),
        content=_xls_fixture_as_xlsx("oborotka_07.xls"),
    )
    dtype, status, rec, err = parse_attachment(att)
    assert dtype == "turnover_statement"
    assert status == "parsed"
    assert rec["period_hint"] == "2026-07"
    assert rec["rows"][0]["employee"] == "ИВАНОВА И.И."
    assert rec["rows"][0]["ndfl"] == "6318.00"
    assert rec["contributions_total"] == "13595.93"
    assert err is None


def test_unknown_xlsx_needs_review_not_error() -> None:
    """Произвольный .xlsx (опись, не ведомость) — «нужна проверка», а не ошибка разбора."""
    from dataclasses import replace

    content = _xlsx_bytes(
        [["Опись документов"], ["Договор аренды", 1], ["Пояснение", 2]]
    )
    att = replace(
        _att("oborotka_07.xls", "Документы в ответ на Требование №1161 от 19.03.26.xlsx"),
        content=content,
    )
    dtype, status, _, err = parse_attachment(att)
    assert dtype == "payroll_statement"
    assert status == "needs_review"
    assert err is None


def test_corrupt_docx_error_is_russian() -> None:
    """Битый .docx падает по-русски с подсказкой, а не «File is not a zip file»."""
    from dataclasses import replace

    att = replace(
        _att("usn_h1_478376.docx", "УСН 2 кв до 28.07.docx"),
        content=b"\x00\x01 not a real docx",
    )
    dtype, status, _, err = parse_attachment(att)
    assert dtype == "unknown"
    assert status == "error"
    assert err is not None
    assert "zip file" not in err
    assert "повреждён" in err


# ── приём в staging ──────────────────────────────────────────────────────────


async def test_ingest_stores_and_routes(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    attachments = [
        _att("usn_h1_478376.docx", "УСН 2 кв до 28.07.docx"),
        _att("enp_payroll_14902.docx", "ЕНП_до 28.07.docx"),
        _att("vedomost_advance_20986.xls", "ВЕД-13 АВАНС 20.07.xls"),
    ]
    async with async_session_factory() as session:
        result = await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub(attachments),
            accounts=_TEST_ACCOUNTS,
        )

    assert result["fetched"] == 3
    assert result["parsed"] == 3  # УСН + ЕНП + ведомость (ЕНП теперь распознан, разнос из оборотки)
    assert result["needs_review"] == 0
    async with async_session_factory() as session:
        assert await session.scalar(
            select(func.count()).select_from(TaxDocumentIntake)
        ) == 3


async def test_ingest_dedups_by_content(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """Повторный проход того же вложения не создаёт вторую строку."""
    att = _att("usn_h1_478376.docx", "УСН 2 кв до 28.07.docx")
    async with async_session_factory() as session:
        await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([att]),
            accounts=_TEST_ACCOUNTS,
        )
    async with async_session_factory() as session:
        result = await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([att]),
            accounts=_TEST_ACCOUNTS,
        )

    assert result["duplicate"] == 1
    assert result["parsed"] == 0
    async with async_session_factory() as session:
        assert await session.scalar(
            select(func.count()).select_from(TaxDocumentIntake)
        ) == 1


async def test_ingest_skips_foreign_sender(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """Письмо не от налогового агента в налоговый staging не попадает."""
    att = _att(
        "usn_h1_478376.docx", "какой-то счёт.docx", sender="supplier@example.com"
    )
    async with async_session_factory() as session:
        result = await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([att]),
            accounts=_TEST_ACCOUNTS,
        )

    assert result["skipped_sender"] == 1
    assert result["parsed"] == 0
    async with async_session_factory() as session:
        assert await session.scalar(
            select(func.count()).select_from(TaxDocumentIntake)
        ) == 0


async def test_ingest_ignores_same_document_in_another_format(
    async_session_factory: async_sessionmaker[AsyncSession],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Ведомость, присланная сначала .xls, а потом .pdf, во второй раз уходит в «отклонён».

    Байты разные — дедуп по SHA-256 её не ловит, а в списке документов возникает призрачный
    дубль (ВЕД-14 пришла .xls 22.07.2026 и .pdf 27.07.2026). Сравниваем содержание.
    """
    from app.services.taxes import document_parser

    xls = _att("vedomost_advance_20986.xls", "ВЕД-13 АВАНС 20.07.xls")
    async with async_session_factory() as session:
        await ingest_tax_documents(
            session, settings=get_settings(), fetch=_fetch_stub([xls]), accounts=_TEST_ACCOUNTS
        )

    _, _, xls_rec, _ = parse_attachment(xls)
    monkeypatch.setattr(
        document_parser,
        "_pdf_text",
        lambda data: "\n".join(
            [
                f"1 {xls_rec['rows'][0]['tab_number']} {xls_rec['rows'][0]['employee']} 20986.00",
                f"{xls_rec['doc_number']} 01.07.2026 31.07.2026",
            ]
        ),
    )
    async with async_session_factory() as session:
        result = await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([_pdf_att("ВЕД-13 АВАНС 20.07.pdf")]),
            accounts=_TEST_ACCOUNTS,
        )

    assert result["ignored"] == 1
    assert result["parsed"] == 0
    async with async_session_factory() as session:
        rows = (
            await session.execute(
                select(TaxDocumentIntake.filename, TaxDocumentIntake.status).order_by(
                    TaxDocumentIntake.filename
                )
            )
        ).all()
    assert [(name, status) for name, status in rows] == [
        ("ВЕД-13 АВАНС 20.07.pdf", "ignored"),
        ("ВЕД-13 АВАНС 20.07.xls", "parsed"),
    ]


async def test_corrected_document_is_not_treated_as_duplicate(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """Исправленный документ (другие суммы) — новый документ, а не дубль."""
    async with async_session_factory() as session:
        await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([_att("vedomost_advance_20986.xls", "ВЕД-13 АВАНС 20.07.xls")]),
            accounts=_TEST_ACCOUNTS,
        )
    async with async_session_factory() as session:
        result = await ingest_tax_documents(
            session,
            settings=get_settings(),
            fetch=_fetch_stub([_att("vedomost_salary_22696.xls", "ВЕД-14 ЗП 05.08.xls")]),
            accounts=_TEST_ACCOUNTS,
        )

    assert result["parsed"] == 1
    assert result["ignored"] == 0


def test_content_sha_is_stable() -> None:
    """Дедуп опирается на SHA содержимого — он детерминирован."""
    att = _att("usn_h1_478376.docx", "УСН 2 кв до 28.07.docx")
    assert att.sha256 == hashlib.sha256(att.content).hexdigest()


# ── ручная проверка документа владельцем ──────────────────────────────────────


def _needs_review_intake(status: str = "needs_review") -> TaxDocumentIntake:
    return TaxDocumentIntake(
        id=uuid.uuid4(),
        mailbox="corporate",
        attachment_sha256=(uuid.uuid4().hex + uuid.uuid4().hex),
        filename="ЕНП_до 28.07.docx",
        document_type="payment_order",
        status=status,
        recognition={},
    )


async def test_review_marks_parsed_then_ignored(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    async with async_session_factory() as session:
        intake = _needs_review_intake()
        session.add(intake)
        await session.flush()

        await set_intake_review(session, intake, status="parsed")
        assert intake.status == "parsed"  # проверено → готово к продвижению

        await set_intake_review(session, intake, status="ignored")
        assert intake.status == "ignored"  # передумал → отклонено


async def test_review_rejects_promoted_and_bad_status(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    async with async_session_factory() as session:
        promoted = _needs_review_intake(status="promoted")
        session.add(promoted)
        await session.flush()
        with pytest.raises(ValueError, match="продвинут"):
            await set_intake_review(session, promoted, status="parsed")

        intake = _needs_review_intake()
        with pytest.raises(ValueError, match="статус"):
            await set_intake_review(session, intake, status="deleted")


async def test_review_with_overrides_fills_missing_fields(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """Проверка с дозаправкой: владелец указывает срок/сумму/вид — они попадают в recognition,
    причины «нужна проверка» снимаются, и продвижение возьмёт исправленные значения."""
    async with async_session_factory() as session:
        intake = _needs_review_intake()
        intake.recognition = {
            "tax_kind": None,
            "amount": "116360",
            "review_reasons": ["не распознан срок уплаты"],
        }
        session.add(intake)
        await session.flush()

        await set_intake_review(
            session,
            intake,
            status="parsed",
            overrides={"tax_kind": "contrib_extra_1pct", "due_date": "2026-04-28"},
        )

        assert intake.status == "parsed"
        assert intake.recognition["tax_kind"] == "contrib_extra_1pct"
        assert intake.recognition["due_date"] == "2026-04-28"
        assert intake.recognition["amount"] == "116360"  # нетронутое осталось
        assert intake.recognition["review_reasons"] == []
        assert intake.recognition["manually_reviewed"] is True


async def test_review_overrides_reject_unknown_kind(
    async_session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """Мусорный вид платежа в ручной правке отклоняется, а не пишется в расчёт."""
    async with async_session_factory() as session:
        intake = _needs_review_intake()
        session.add(intake)
        await session.flush()

        with pytest.raises(ValueError, match="Неизвестный вид"):
            await set_intake_review(
                session, intake, status="parsed", overrides={"tax_kind": "чепуха"}
            )


def test_svod_classified_as_turnover() -> None:
    """Файл «СВОД 02.xls» — это оборотка в другой форме, а не ведомость Т-53.

    До правки свод уходил в разбор ведомостей и вис в «нужна проверка»: взносы февраля
    в канон не попадали.
    """
    from app.services.taxes.document_ingest import _classify_document

    assert _classify_document("СВОД 02.xls") == "turnover_statement"
    assert _classify_document("ОБОРОТКА 01.xls") == "turnover_statement"
    # «Сводная ведомость» тоже свод; а «ВЕД-3 АВАНС» остаётся ведомостью Т-53.
    assert _classify_document("Сводная 03.xls") == "turnover_statement"
    assert _classify_document("ВЕД-3 АВАНС 20.02.xls") == "payroll_statement"
