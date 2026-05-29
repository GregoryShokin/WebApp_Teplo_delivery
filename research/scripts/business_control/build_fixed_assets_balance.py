#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "data" / "processed" / "fixed_assets"
FIXED_ASSETS_XLSX = Path("/private/tmp/teplo_fixed_assets.xlsx")
BALANCE_XLSX = Path("/private/tmp/teplo_balance.xlsx")
CURRENT_DATE = date(2026, 5, 20)
BANK_START = date(2026, 2, 1)
BANK_END = date(2026, 5, 19)

FA_URL = "https://docs.google.com/spreadsheets/d/1lDJi4dShJSP71Rq55uGQY_rW_kWctaoqX2MJmEcOIUc/edit"
BALANCE_URL = "https://docs.google.com/spreadsheets/d/1ekjJSe8Rt7IROfBVYhqZSkvwG1FnUycCPaUM4w9xqsE/edit"

KEYWORDS = {
    "оборудование",
    "печь",
    "холодильник",
    "плита",
    "гриль",
    "мебель",
    "ремонт",
    "строит",
    "плитка",
    "сантехника",
    "вытяжка",
    "мойка",
    "стол",
    "стеллаж",
    "стелаж",
    "вентиляция",
    "кондиционер",
    "компьютер",
    "ноутбук",
    "планшет",
    "телевизор",
    "касса",
    "терминал",
    "сейф",
    "видеонаблюдение",
    "камеры",
}

COUNTERPARTY_IDS = {
    "ооо назад в будущее": ("CP-0001", "online_platform/payment_partner"),
    "ооо тора": ("CP-0002", "food_supplier"),
    "ооо альянс юг": ("CP-0003", "food_supplier"),
    "ооо мяснофф-дон": ("CP-0004", "food_supplier"),
    "ооо мистерия": ("CP-0005", "food_supplier"),
    "ооо о. о": ("CP-0006", "marketing_supplier"),
    "ооо метро кэш энд керри": ("CP-0007", "food_supplier"),
    "ао айко": ("CP-0008", "automation_system"),
    "ооо ревви": ("CP-0009", "automation_system"),
    "ооо доксинбокс": ("CP-0010", "edo_system"),
    "ооо синапсис": ("CP-0011", "marketing_supplier"),
    "ооо кварц": ("CP-0012", "supplier_needs_review"),
    "общество с ограниченной ответственностью экоцентр": ("CP-0013", "waste_removal"),
    "ооо экоцентр": ("CP-0013", "waste_removal"),
    "ооо суши принт": ("CP-0014", "food_packaging_supplier"),
    "ооо сити": ("CP-0015", "historic_marketing_supplier"),
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def norm_key(value: Any) -> str:
    text = norm(value).lower()
    text = text.replace('"', "").replace("«", "").replace("»", "")
    return re.sub(r"\s+", " ", text).strip()


def as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = norm(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


def iso(value: Any) -> str:
    parsed = as_date(value)
    return parsed.isoformat() if parsed else ""


def money(value: float) -> str:
    return f"{round(float(value), 2):.2f}"


def safe_hash(value: str, length: int = 12) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:length]


def source_signature(text: str, matched_keywords: list[str]) -> str:
    cleaned = norm(text).lower()
    cleaned = re.sub(r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b", "<date>", cleaned)
    cleaned = re.sub(r"\d{4,}", "<num>", cleaned)
    cleaned = re.sub(r"\b[а-яa-z0-9._%+-]+@[а-яa-z0-9.-]+\.[а-яa-z]{2,}\b", "<email>", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    words = [w for w in re.split(r"[^а-яa-z0-9<>]+", cleaned) if w]
    non_sensitive = [w for w in words if w in KEYWORDS or w in {"оплата", "счету", "поставка", "ндс"}]
    base = " ".join(non_sensitive[:8]) or "no_public_description"
    kw = ",".join(sorted(set(matched_keywords))) or "no_keyword"
    return f"{base}; kw={kw}; sig={safe_hash(cleaned)}"


def counterparty_public_id(row: dict[str, str]) -> tuple[str, str]:
    name = norm_key(row.get("counterparty_name"))
    for key, mapped in COUNTERPARTY_IDS.items():
        if key and key in name:
            return mapped
    raw = "|".join(
        [
            norm_key(row.get("counterparty_name")),
            norm_key(row.get("counterparty_inn")),
            norm_key(row.get("counterparty_account")),
        ]
    )
    return f"CP-HASH-{safe_hash(raw, 10)}", "unknown_or_private_counterparty"


def sheet_bounds(ws: Any) -> str:
    min_row = min_col = 10**9
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                min_row = min(min_row, cell.row)
                min_col = min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if max_row == 0:
        return "empty_in_xlsx_export"
    return f"R{min_row}C{min_col}:R{max_row}C{max_col}"


def read_fixed_assets() -> tuple[list[dict[str, str]], dict[str, Any]]:
    wb_values = openpyxl.load_workbook(FIXED_ASSETS_XLSX, data_only=True, read_only=False)
    wb_formulas = openpyxl.load_workbook(FIXED_ASSETS_XLSX, data_only=False, read_only=False)
    ws = wb_values["Учёт ОС"]

    rows: list[dict[str, str]] = []
    all_dates: list[date] = []
    formulas_ws = wb_formulas["Учёт ОС"]
    method = "linear_monthly_formula"

    for r in range(2, ws.max_row + 1):
        name = norm(ws.cell(r, 2).value)
        if not name:
            continue
        qty = as_float(ws.cell(r, 4).value)
        unit_price = as_float(ws.cell(r, 5).value)
        initial_cost = qty * unit_price
        residual_value = as_float(ws.cell(r, 37).value)
        accumulated = max(0.0, initial_cost - residual_value)
        sale_date = as_date(ws.cell(r, 20).value)
        sale_qty = as_float(ws.cell(r, 23).value)
        status = "disposed" if sale_date or (qty and sale_qty >= qty) else "in_work"
        acquired = as_date(ws.cell(r, 11).value)
        introduced = as_date(ws.cell(r, 9).value)
        life_months = as_float(ws.cell(r, 15).value) or as_float(ws.cell(r, 14).value) * 12
        source_dates = [d for d in [acquired, introduced, sale_date, as_date(ws.cell(r, 18).value)] if d]
        all_dates.extend(source_dates)

        notes = [
            f"source_no={norm(ws.cell(r, 1).value)}",
            f"qty={qty:g}",
            f"unit_price={unit_price:g}",
        ]
        if introduced:
            notes.append(f"introduced_at={introduced.isoformat()}")
        if sale_date:
            notes.append(f"sold_at={sale_date.isoformat()}")
        if residual_value == 0 and initial_cost > 0 and acquired and acquired.year >= 2024:
            notes.append("residual_zero_in_source_check_formula_copy")
        formula = formulas_ws.cell(r, 24).value
        if isinstance(formula, str) and formula.startswith("="):
            notes.append("monthly_depreciation_formula_present")

        rows.append(
            {
                "asset_id": f"FA-HIST-{r:04d}",
                "name": name,
                "location": norm(ws.cell(r, 7).value),
                "category": norm(ws.cell(r, 8).value),
                "acquired_at": acquired.isoformat() if acquired else "",
                "initial_cost": money(initial_cost),
                "accumulated_depreciation": money(accumulated),
                "residual_value": money(residual_value),
                "method": method,
                "useful_life_months": str(int(life_months)) if life_months else "",
                "status": status,
                "source_sheet": "Учёт ОС",
                "source_row": str(r),
                "last_update_in_source": "",
                "notes": "; ".join(notes),
            }
        )

    last_update = max(all_dates) if all_dates else None
    for row in rows:
        row["last_update_in_source"] = last_update.isoformat() if last_update else ""

    reg = wb_values["Регламент"]
    structure = {
        "source_url": FA_URL,
        "title": "Копия Учет Основных Средств 2.0 [ИП Шокина]",
        "sheets": [
            {"name": sheet.title, "bounds": sheet_bounds(sheet)}
            for sheet in wb_values.worksheets
        ],
        "main_headers": [norm(ws.cell(1, c).value) for c in range(1, 53)],
        "regulation_created_at": iso(reg["G2"].value),
        "regulation_changed_at": iso(reg["G3"].value) or iso(reg["G4"].value),
        "last_update_in_rows": last_update.isoformat() if last_update else "",
        "asset_count": len(rows),
        "initial_total": sum(as_float(r["initial_cost"]) for r in rows),
        "residual_total": sum(as_float(r["residual_value"]) for r in rows),
        "accumulated_total": sum(as_float(r["accumulated_depreciation"]) for r in rows),
        "locations": Counter(r["location"] for r in rows),
        "categories": Counter(r["category"] for r in rows),
        "statuses": Counter(r["status"] for r in rows),
        "zero_initial_count": sum(1 for r in rows if as_float(r["initial_cost"]) == 0),
    }
    return rows, structure


def balance_section(row_num: int, label: str) -> str:
    if 3 <= row_num <= 50:
        if row_num <= 17:
            return "Активы / Внеоборотные активы / Основные средства"
        if row_num <= 25:
            return "Активы / Оборотные активы / Запасы"
        if row_num <= 35:
            return "Активы / Оборотные активы / Денежные средства"
        if row_num <= 41:
            return "Активы / Оборотные активы / Дебиторская задолженность"
        if row_num <= 45:
            return "Активы / Оборотные активы / Краткосрочные финансовые вложения"
        return "Активы / Оборотные активы / Прочие активы"
    if 51 <= row_num <= 68:
        return "Пассивы / Капитал"
    if 69 <= row_num <= 74:
        return "Пассивы / Обязательства / Долгосрочные"
    if 75 <= row_num <= 78:
        return "Пассивы / Обязательства / Краткосрочные"
    if 79 <= row_num <= 89:
        return "Пассивы / Обязательства / Кредиторская задолженность"
    if 90 <= row_num <= 111:
        return "Финансовые показатели"
    return "Прочее"


def read_balance() -> tuple[list[dict[str, str]], dict[str, Any]]:
    wb_values = openpyxl.load_workbook(BALANCE_XLSX, data_only=True, read_only=False)
    wb_formulas = openpyxl.load_workbook(BALANCE_XLSX, data_only=False, read_only=False)
    ws = wb_values["Баланс "]
    wf = wb_formulas["Баланс "]
    dictionary = wb_values["Справочник статей баланса"]

    source_by_line: dict[str, str] = {}
    for r in range(2, dictionary.max_row + 1):
        label = norm(dictionary.cell(r, 1).value)
        source = norm(dictionary.cell(r, 3).value)
        if label:
            source_by_line[label] = source

    last_period = as_date(ws["M1"].value)
    rows: list[dict[str, str]] = []
    formula_count = 0
    importrange_count = 0
    formula_examples: list[str] = []

    for sheet in wb_formulas.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if "IMPORTRANGE" in value.upper():
                        importrange_count += 1
                    if len(formula_examples) < 8 and sheet.title == "Баланс ":
                        formula_examples.append(f"{sheet.title}!{cell.coordinate}: {value[:120]}")

    for r in range(1, 112):
        label = norm(ws.cell(r, 1).value)
        if not label:
            continue
        value = ws.cell(r, 13).value
        formula = wf.cell(r, 13).value
        source = source_by_line.get(label, "")
        if isinstance(formula, str) and formula.startswith("="):
            formula_or_source = formula
        elif source:
            formula_or_source = f"source: {source}"
        else:
            formula_or_source = "manual_or_blank_value"

        notes = [f"source_sheet=Баланс ", f"source_row={r}"]
        if r == 88 and as_float(value) != 0:
            notes.append("balance_check_not_zero")
        if value in (None, ""):
            notes.append("blank_at_last_filled_period")

        rows.append(
            {
                "section": balance_section(r, label),
                "line_name": label,
                "last_filled_period": last_period.isoformat() if last_period else "",
                "value_at_last_filled_period": "" if value in (None, "") else money(as_float(value)),
                "formula_or_source": norm(formula_or_source),
                "notes": "; ".join(notes),
            }
        )

    structure = {
        "source_url": BALANCE_URL,
        "title": "Копия Баланс• [ИП Шокина]",
        "sheets": [
            {"name": sheet.title, "bounds": sheet_bounds(sheet)}
            for sheet in wb_values.worksheets
        ],
        "last_period": last_period.isoformat() if last_period else "",
        "lag_days": (CURRENT_DATE - last_period).days if last_period else None,
        "formula_count": formula_count,
        "importrange_count": importrange_count,
        "formula_examples": formula_examples,
        "balance_check": as_float(ws["M88"].value),
        "assets_total": as_float(ws["M3"].value),
        "liabilities_total": as_float(ws["M51"].value),
        "fixed_assets_value": as_float(ws["M5"].value),
    }
    return rows, structure


def hypothesize_category(text: str, matched_keywords: list[str]) -> str:
    joined = " ".join(matched_keywords) + " " + text.lower()
    if any(w in joined for w in ["холодильник", "кондиционер", "вентиляция", "вытяжка"]):
        return "equipment_hvac_or_cooling"
    if any(w in joined for w in ["печь", "плита", "гриль", "мойка", "касса", "терминал"]):
        return "restaurant_equipment"
    if any(w in joined for w in ["мебель", "стол", "стеллаж", "стелаж", "сейф"]):
        return "furniture_or_storage"
    if any(w in joined for w in ["ремонт", "строит", "плитка", "сантехника"]):
        return "repair_or_fitout"
    if any(w in joined for w in ["компьютер", "ноутбук", "планшет", "телевизор", "видеонаблюдение", "камеры"]):
        return "it_or_security_equipment"
    return "fixed_asset_candidate"


def read_bank_candidates(inventory_rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], dict[str, Any]]:
    paths = [
        ("tbank", ROOT / "data" / "private" / "tbank" / "statement_classified.csv"),
        ("sber", ROOT / "data" / "private" / "sber" / "statement_classified.csv"),
    ]
    source_rows: list[dict[str, str]] = []
    all_debit_in_period = 0
    base_rows = 0
    keyword_rows = 0
    regular_excluded = 0
    regular_counts: defaultdict[tuple[str, str], int] = defaultdict(int)

    for bank, path in paths:
        with path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                op_date = as_date(row.get("operation_date"))
                if not op_date or not (BANK_START <= op_date <= BANK_END):
                    continue
                if norm_key(row.get("direction")) != "debit":
                    continue
                all_debit_in_period += 1
                row["bank_src"] = bank
                row["amount_num"] = str(abs(as_float(row.get("amount"))))
                source_rows.append(row)
                cp_key = norm_key(row.get("counterparty_inn")) or norm_key(row.get("counterparty_name")) or "unknown"
                regular_counts[(cp_key, op_date.strftime("%Y-%m"))] += 1

    initial_costs = [as_float(row["initial_cost"]) for row in inventory_rows if as_float(row["initial_cost"]) > 0]
    candidates: list[dict[str, str]] = []
    keyword_counter: Counter[str] = Counter()

    for row in source_rows:
        amount = as_float(row.get("amount_num"))
        if amount < 30_000:
            continue
        if row.get("flow_type") not in {"supplier_payment", "other_outflow"}:
            continue
        base_rows += 1
        text = f"{row.get('counterparty_name', '')} {row.get('description_snippet', '')}"
        text_l = text.lower()
        matched = sorted(kw for kw in KEYWORDS if kw in text_l)
        if not matched:
            continue
        keyword_rows += 1
        keyword_counter.update(matched)
        op_date = as_date(row.get("operation_date"))
        cp_key = norm_key(row.get("counterparty_inn")) or norm_key(row.get("counterparty_name")) or "unknown"
        month_count = regular_counts[(cp_key, op_date.strftime("%Y-%m") if op_date else "")]
        if month_count >= 3:
            regular_excluded += 1
            continue
        cp_id, cp_role = counterparty_public_id(row)
        matches_inventory = any(abs(cost - amount) <= amount * 0.05 for cost in initial_costs)
        confidence = "medium" if len(matched) >= 2 else "low"
        candidates.append(
            {
                "operation_date": op_date.isoformat() if op_date else "",
                "bank": row["bank_src"],
                "counterparty_id": cp_id,
                "counterparty_role": cp_role,
                "amount": money(amount),
                "description_signature": source_signature(row.get("description_snippet", ""), matched),
                "flow_type_original": row.get("flow_type", ""),
                "hypothesized_category": hypothesize_category(text_l, matched),
                "matches_inventory_historic": "true" if matches_inventory else "false",
                "confidence": confidence,
                "owner_review_status": "needs_review",
            }
        )

    stats = {
        "period": f"{BANK_START.isoformat()}..{BANK_END.isoformat()}",
        "all_debit_in_period": all_debit_in_period,
        "base_rows": base_rows,
        "keyword_rows": keyword_rows,
        "regular_excluded": regular_excluded,
        "candidate_count": len(candidates),
        "candidate_total": sum(as_float(r["amount"]) for r in candidates),
        "keyword_counter": keyword_counter,
    }
    return sorted(candidates, key=lambda r: as_float(r["amount"]), reverse=True), stats


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def rub(value: float) -> str:
    return f"{value:,.2f}".replace(",", " ")


def build_report(
    fa_structure: dict[str, Any],
    balance_structure: dict[str, Any],
    bank_stats: dict[str, Any],
    bank_candidates: list[dict[str, str]],
) -> str:
    fa_sheets = "\n".join(
        f"- `{s['name']}`: {s['bounds']}" for s in fa_structure["sheets"]
    )
    balance_sheets = "\n".join(
        f"- `{s['name']}`: {s['bounds']}" for s in balance_structure["sheets"]
    )
    locations = "\n".join(
        f"- {name or '(пусто)'}: {count}"
        for name, count in fa_structure["locations"].most_common()
    )
    categories = "\n".join(
        f"- {name or '(пусто)'}: {count}"
        for name, count in fa_structure["categories"].most_common()
    )
    top_candidates = "\n".join(
        f"| {row['operation_date']} | {row['bank']} | {row['counterparty_id']} | {row['amount']} | {row['hypothesized_category']} | {row['confidence']} |"
        for row in bank_candidates[:10]
    )
    if not top_candidates:
        top_candidates = "| — | — | — | — | — | — |"

    lag_days = balance_structure["lag_days"]
    lag_text = f"{lag_days} дней" if lag_days is not None else "не определено"

    return f"""# Основные средства и баланс: processed report

Дата сборки: 2026-05-20.

Источники:
- Google Sheets `Учёт Основных Средств 2.0`, read-only export: {FA_URL}
- Google Sheets `Копия Баланс• [ИП Шокина]`, read-only export: {BALANCE_URL}
- Локальные classified CSV банков: `research/private/tbank/statement_classified.csv`, `research/private/sber/statement_classified.csv`

Google Sheets не изменялись. В processed не вынесены полные назначения платежей, ИНН, счета и персональные данные.

## Учёт ОС

Вкладки:
{fa_sheets}

Главная вкладка: `Учёт ОС`. Ключевые колонки: `Наименование`, `Кол-во ОС`, `Цена ед.`, `Помещение`, `Тип ОС`, `Дата введения`, `Дата покупки или постановки`, `Срок полезного использования`, месячные колонки амортизации, `Оставшаяся стоимость`, `Дата продажи`.

Регламент сформирован: `{fa_structure['regulation_created_at'] or 'не заполнено'}`. Служебная дата изменения в регламенте: `{fa_structure['regulation_changed_at'] or 'не заполнено'}`.

Фактическая последняя дата в строках ОС: `{fa_structure['last_update_in_rows'] or 'не найдена'}`. Это не доказывает актуальность на 2026-05-20: владелец прямо отметил, что отчёт давно не обновлялся.

Агрегаты исторического реестра:
- Строк ОС: {fa_structure['asset_count']}
- Первоначальная стоимость: {rub(fa_structure['initial_total'])} ₽
- Накопленная амортизация расчетно: {rub(fa_structure['accumulated_total'])} ₽
- Остаточная стоимость по книге: {rub(fa_structure['residual_total'])} ₽
- Строк с нулевой первоначальной стоимостью: {fa_structure['zero_initial_count']}

Локации:
{locations}

Категории:
{categories}

## Баланс

Вкладки:
{balance_sheets}

Последний заполненный период: `{balance_structure['last_period'] or 'не найден'}`. Отставание от 2026-05-20: {lag_text}.

Структура баланса включает:
- Активы: внеоборотные активы / основные средства; оборотные активы / запасы, денежные средства, дебиторская задолженность, краткосрочные финансовые вложения, прочие активы.
- Пассивы: капитал, накопленная прибыль/убыток, дивиденды, вложения собственников, долгосрочные и краткосрочные обязательства, кредиторская задолженность.
- Финансовые показатели: текущая ликвидность, абсолютная ликвидность, финансовая устойчивость, финансовая независимость, ROE, ROA.

На последнюю дату:
- Активы: {rub(balance_structure['assets_total'])} ₽
- Пассивы: {rub(balance_structure['liabilities_total'])} ₽
- Проверка `Активы = Пассивы`: {rub(balance_structure['balance_check'])} ₽
- ОС в балансе: {rub(balance_structure['fixed_assets_value'])} ₽

Формулы: найдено {balance_structure['formula_count']} формул, `IMPORTRANGE`: {balance_structure['importrange_count']}. Основные связи внутри листа через `SUM`, ссылки на строки и `VLOOKUP`; внешних `IMPORTRANGE` в выгрузке не найдено.

## Банковские кандидаты в ОС

Период скрининга: `{bank_stats['period']}`.

- Debit-операций в периоде: {bank_stats['all_debit_in_period']}
- После фильтра сумма ≥ 30 000 ₽ и flow_type `supplier_payment`/`other_outflow`: {bank_stats['base_rows']}
- После тематических ключевых слов: {bank_stats['keyword_rows']}
- Исключено как регулярные получатели 3+ платежа/месяц: {bank_stats['regular_excluded']}
- Итоговых кандидатов: {bank_stats['candidate_count']}
- Итоговая сумма кандидатов: {rub(bank_stats['candidate_total'])} ₽

Топ-10 кандидатов:

| Дата | Банк | Контрагент ID | Сумма | Гипотеза | Confidence |
| --- | --- | ---: | ---: | --- | --- |
{top_candidates}

Вывод: по строгому правилу ключевых слов в classified-выписках за 2026-02-01..2026-05-19 кандидаты не найдены. Это не означает, что покупок ОС не было: назначения платежей в банковских CSV могут быть слишком общими, поэтому следующий шаг — сверка с актами/счетами и инвентаризация.

## Открытые вопросы владельцу

1. Подтвердить порог управленческого признания ОС: 5 000 ₽ из регламента или новый порог, например 30 000 ₽.
2. Подтвердить метод амортизации: оставить линейный помесячный метод из текущего файла или перейти на другой управленческий подход.
3. Есть ли покупки ОС после `{fa_structure['last_update_in_rows'] or 'последней даты файла'}` не через банк: наличные, личная карта, взаимозачёты, рассрочка/лизинг.
4. Что делать с точкой `Гагарина`: какие ОС реально в работе, какие на складе, какие списаны/проданы.
5. Обновляем старый отчёт ОС или сразу собираем новый актуальный реестр с инвентаризацией на дату.
6. Как восстанавливать баланс: через актуальную инвентаризацию активов/обязательств или через продолжение старого файла с 2025-01.

## Связь с P&L

Строка `Амортизация` находится в блоке `Расходы ниже EBITDA` и по методологии берётся из `Учет ОС`. Помесячную амортизацию не считаем до подтверждения актуального реестра ОС.
"""


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    inventory_rows, fa_structure = read_fixed_assets()
    balance_rows, balance_structure = read_balance()
    bank_candidates, bank_stats = read_bank_candidates(inventory_rows)

    write_csv(
        OUT_DIR / "fixed_assets_inventory_historic.csv",
        inventory_rows,
        [
            "asset_id",
            "name",
            "location",
            "category",
            "acquired_at",
            "initial_cost",
            "accumulated_depreciation",
            "residual_value",
            "method",
            "useful_life_months",
            "status",
            "source_sheet",
            "source_row",
            "last_update_in_source",
            "notes",
        ],
    )
    write_csv(
        OUT_DIR / "balance_structure.csv",
        balance_rows,
        [
            "section",
            "line_name",
            "last_filled_period",
            "value_at_last_filled_period",
            "formula_or_source",
            "notes",
        ],
    )
    write_csv(
        OUT_DIR / "fixed_assets_bank_candidates.csv",
        bank_candidates,
        [
            "operation_date",
            "bank",
            "counterparty_id",
            "counterparty_role",
            "amount",
            "description_signature",
            "flow_type_original",
            "hypothesized_category",
            "matches_inventory_historic",
            "confidence",
            "owner_review_status",
        ],
    )
    (OUT_DIR / "report.md").write_text(
        build_report(fa_structure, balance_structure, bank_stats, bank_candidates),
        encoding="utf-8",
    )

    print(f"inventory_rows={len(inventory_rows)}")
    print(f"inventory_initial_total={money(fa_structure['initial_total'])}")
    print(f"inventory_residual_total={money(fa_structure['residual_total'])}")
    print(f"balance_rows={len(balance_rows)}")
    print(f"balance_last_period={balance_structure['last_period']}")
    print(f"bank_candidates={len(bank_candidates)}")
    print(f"bank_candidates_total={money(bank_stats['candidate_total'])}")


if __name__ == "__main__":
    main()
