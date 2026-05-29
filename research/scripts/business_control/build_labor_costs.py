#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import math
import re
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


FOCUS_MONTHS = ["2026-02", "2026-03", "2026-04", "2026-05"]
COMPARISON_MONTHS = ["2025-11", "2025-12", "2026-01"] + FOCUS_MONTHS

PAYROLL_EARNING_TYPES = {
    "оклад",
    "процент",
    "премия",
    "накопительный фонд",
    "больничные и отпуска и пособия",
}
PAYROLL_DEDUCTION_TYPES = {
    "штрафы по ревизиям",
    "штрафы и удержания",
    "депозит удержание",
    "ндфл начислено",
    "депозит списание",
    "списание накоплений",
}
PAYROLL_OTHER_TYPES = {
    "депозит возврат",
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_key(value: Any) -> str:
    return re.sub(r"\s+", " ", norm(value).lower())


def as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = norm(value)
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            pass
    return None


def month_key(value: date | datetime | str | None) -> str | None:
    d = as_date(value)
    return d.strftime("%Y-%m") if d else None


def round_money(value: float) -> float:
    return round(float(value), 2)


def pct(numerator: float, denominator: float) -> float | str:
    if not denominator:
        return ""
    return round(numerator / denominator, 6)


def role_bucket(role: str, unit: str) -> str:
    role_l = norm_key(role)
    unit_l = norm_key(unit)
    if "администрация" in unit_l:
        return "central_admin"
    if "администратор" in role_l:
        return "shift_admin"
    if "сушист" in role_l:
        return "sushi"
    if "пицц" in role_l:
        return "pizza"
    if "шаурм" in role_l:
        return "shawarma"
    if "заготов" in role_l:
        return "prep"
    if "производство" in unit_l:
        return "other_production"
    return "other"


def read_sales_monthly(path: Path) -> dict[str, dict[str, float]]:
    out: dict[str, dict[str, float]] = defaultdict(lambda: {"orders": 0.0, "revenue": 0.0, "gross_sum": 0.0})
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            m = row["date"][:7]
            if m in COMPARISON_MONTHS:
                out[m]["orders"] += as_float(row.get("orders"))
                out[m]["revenue"] += as_float(row.get("revenue"))
                out[m]["gross_sum"] += as_float(row.get("gross_sum"))
    return out


def read_delivery_monthly(raw_dir: Path) -> dict[str, dict[str, float]]:
    out: dict[str, dict[str, float]] = defaultdict(lambda: {"delivery_orders_iiko": 0.0, "delivery_revenue_iiko": 0.0})
    for path in sorted(raw_dir.glob("olap_channels_*.xml")):
        m = re.search(r"_(\d{4})-(\d{2})-\d{2}_", path.name)
        if not m:
            continue
        month = f"{m.group(1)}-{m.group(2)}"
        if month not in COMPARISON_MONTHS:
            continue
        try:
            root = ET.parse(path).getroot()
        except ET.ParseError:
            continue
        for r in root.findall(".//r"):
            service_type = norm(r.findtext("Delivery.ServiceType")) or "(пусто)"
            if service_type != "COURIER":
                continue
            out[month]["delivery_orders_iiko"] += as_float(r.findtext("OrderNum"))
            out[month]["delivery_revenue_iiko"] += as_float(r.findtext("DishDiscountSumInt"))
    return out


@dataclass
class PayrollMonth:
    accrual_total: float = 0.0
    accrual_total_including_deductions: float = 0.0
    paid_total: float = 0.0
    rows_accrual: int = 0
    rows_paid: int = 0
    by_role: Counter = field(default_factory=Counter)
    by_type: Counter = field(default_factory=Counter)
    deductions: Counter = field(default_factory=Counter)
    other_types: Counter = field(default_factory=Counter)
    payment_types: Counter = field(default_factory=Counter)
    filled_dates: set[date] = field(default_factory=set)
    paid_dates: set[date] = field(default_factory=set)


def read_payroll(path: Path) -> tuple[dict[str, PayrollMonth], dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    monthly: dict[str, PayrollMonth] = defaultdict(PayrollMonth)
    type_inventory: Counter = Counter()
    role_inventory: Counter = Counter()

    for row in wb["Выгрузка"].iter_rows(values_only=True):
        d = as_date(row[0] if row else None)
        month = month_key(d)
        if month not in COMPARISON_MONTHS:
            continue
        amount = as_float(row[1] if len(row) > 1 else None)
        role = norm(row[3] if len(row) > 3 else "")
        unit = norm(row[4] if len(row) > 4 else "")
        type_name = norm(row[5] if len(row) > 5 else "")
        type_key = norm_key(type_name)
        bucket = role_bucket(role, unit)

        pm = monthly[month]
        pm.rows_accrual += 1
        pm.filled_dates.add(d)
        pm.accrual_total_including_deductions += amount
        type_inventory[type_name] += 1
        role_inventory[role] += 1

        if type_key in PAYROLL_EARNING_TYPES:
            pm.accrual_total += amount
            pm.by_role[bucket] += amount
            pm.by_type[type_name] += amount
        elif type_key in PAYROLL_DEDUCTION_TYPES:
            pm.deductions[type_name] += amount
        elif type_key in PAYROLL_OTHER_TYPES:
            pm.other_types[type_name] += amount
        elif type_key:
            pm.other_types[type_name] += amount

    for row in wb["Выплаты"].iter_rows(values_only=True):
        d = as_date(row[0] if row else None)
        month = month_key(d)
        if month not in COMPARISON_MONTHS:
            continue
        amount = as_float(row[1] if len(row) > 1 else None)
        role = norm(row[3] if len(row) > 3 else "")
        unit = norm(row[4] if len(row) > 4 else "")
        payment_type = norm(row[5] if len(row) > 5 else "")
        bucket = role_bucket(role, unit)
        pm = monthly[month]
        pm.paid_total += amount
        pm.rows_paid += 1
        pm.paid_dates.add(d)
        pm.payment_types[payment_type] += amount
        # Keep paid role buckets only as diagnostics in source quality; the output stays aggregate.
        pm.other_types[f"paid_role::{bucket}"] += amount

    admin_sheet_sample = []
    for i, row in enumerate(wb["Зарплата Администрации"].iter_rows(values_only=True), start=1):
        if i > 8:
            break
        admin_sheet_sample.append([norm(v) for v in row[:7]])

    structure = {
        "sheets": wb.sheetnames,
        "payroll_type_inventory": dict(type_inventory),
        "role_inventory_count": len(role_inventory),
        "admin_sheet_current_sample_rows": len(admin_sheet_sample),
    }
    wb.close()
    return monthly, structure


@dataclass
class CourierMonth:
    delivery_rows: int = 0
    unique_delivery_orders: set[str] = field(default_factory=set)
    completed_delivery_rows: int = 0
    delivery_duration_hours_sum: float = 0.0
    shift_rows: int = 0
    completed_shift_rows: int = 0
    shift_hours_sum: float = 0.0
    help_shift_rows: int = 0
    schedule_status: Counter = field(default_factory=Counter)
    dates_delivery: set[date] = field(default_factory=set)
    dates_shift: set[date] = field(default_factory=set)


def read_couriers(path: Path) -> tuple[dict[str, CourierMonth], dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    monthly: dict[str, CourierMonth] = defaultdict(CourierMonth)

    for row in wb["Доставки"].iter_rows(values_only=True):
        d = as_date(row[0] if row else None)
        month = month_key(d)
        if month not in COMPARISON_MONTHS:
            continue
        cm = monthly[month]
        cm.delivery_rows += 1
        cm.dates_delivery.add(d)
        order_id = norm(row[5] if len(row) > 5 else "") or norm(row[4] if len(row) > 4 else "")
        if order_id:
            cm.unique_delivery_orders.add(order_id)
        duration = as_float(row[6] if len(row) > 6 else None)
        if duration > 0:
            cm.completed_delivery_rows += 1
            cm.delivery_duration_hours_sum += duration

    for row in wb["Выходы"].iter_rows(values_only=True):
        d = as_date(row[0] if row else None)
        month = month_key(d)
        if month not in COMPARISON_MONTHS:
            continue
        cm = monthly[month]
        cm.shift_rows += 1
        cm.dates_shift.add(d)
        hours = as_float(row[3] if len(row) > 3 else None)
        if hours > 0:
            cm.completed_shift_rows += 1
            cm.shift_hours_sum += hours
        if norm_key(row[6] if len(row) > 6 else "") == "помог":
            cm.help_shift_rows += 1

    # The visual schedule currently has February date columns only in the export.
    ws = wb["График"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) >= 2:
        date_row = rows[1]
        for r in rows[2:]:
            if not any(v not in (None, "") for v in r):
                continue
            for idx, cell_date in enumerate(date_row[1:], start=1):
                d = as_date(cell_date)
                month = month_key(d)
                if month not in COMPARISON_MONTHS:
                    continue
                status = norm(r[idx] if idx < len(r) else "")
                if status:
                    monthly[month].schedule_status[status] += 1

    tech_values = []
    for i, row in enumerate(wb["Технический лист"].iter_rows(values_only=True), start=1):
        if i > 28:
            break
        tech_values.append(row[:8])
    structure = {
        "sheets": wb.sheetnames,
        "technical_list_rows_sampled": len(tech_values),
    }
    wb.close()
    return monthly, structure


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def date_span(dates: set[date]) -> str:
    if not dates:
        return ""
    return f"{min(dates).isoformat()}..{max(dates).isoformat()}"


def build_outputs(args: argparse.Namespace) -> None:
    out_dir = Path(args.out_dir)
    sales = read_sales_monthly(Path(args.sales_daily))
    delivery = read_delivery_monthly(Path(args.iiko_orders_raw_dir))
    payroll, payroll_structure = read_payroll(Path(args.payroll_xlsx))
    couriers, courier_structure = read_couriers(Path(args.couriers_xlsx))

    payroll_rows: list[dict[str, Any]] = []
    for month in COMPARISON_MONTHS:
        pm = payroll.get(month, PayrollMonth())
        revenue = sales.get(month, {}).get("revenue", 0.0)
        kitchen = (
            pm.by_role["sushi"]
            + pm.by_role["pizza"]
            + pm.by_role["shawarma"]
            + pm.by_role["prep"]
            + pm.by_role["other_production"]
        )
        payroll_rows.append(
            {
                "month": month,
                "iiko_revenue": round_money(revenue),
                "payroll_accrual_total": round_money(pm.accrual_total),
                "payroll_cash_paid_total": round_money(pm.paid_total),
                "central_admin_accrual": round_money(pm.by_role["central_admin"]),
                "shift_admin_accrual": round_money(pm.by_role["shift_admin"]),
                "kitchen_accrual": round_money(kitchen),
                "sushi_accrual": round_money(pm.by_role["sushi"]),
                "pizza_accrual": round_money(pm.by_role["pizza"]),
                "shawarma_accrual": round_money(pm.by_role["shawarma"]),
                "prep_accrual": round_money(pm.by_role["prep"]),
                "other_production_accrual": round_money(pm.by_role["other_production"]),
                "other_role_accrual": round_money(pm.by_role["other"]),
                "base_shift_accrual": round_money(pm.by_type["Оклад"]),
                "percent_accrual": round_money(pm.by_type["Процент"]),
                "bonus_accrual": round_money(pm.by_type["Премия"]),
                "savings_fund_accrual": round_money(pm.by_type["Накопительный фонд"]),
                "sick_leave_vacation_accrual": round_money(pm.by_type["Больничные и отпуска и пособия"]),
                "fines_revision_amount": round_money(pm.deductions["Штрафы по ревизиям"]),
                "fines_other_amount": round_money(pm.deductions["Штрафы и удержания"]),
                "deposit_withheld_amount": round_money(pm.deductions["Депозит удержание"]),
                "deposit_return_amount": round_money(pm.other_types["Депозит возврат"]),
                "ndfl_accrued_amount": round_money(pm.deductions["НДФЛ Начислено"]),
                "other_non_fot_amount": round_money(
                    sum(v for k, v in pm.other_types.items() if not k.startswith("paid_role::") and k != "Депозит возврат")
                ),
                "payroll_accrual_pct_of_revenue": pct(pm.accrual_total, revenue),
                "payroll_cash_paid_pct_of_revenue": pct(pm.paid_total, revenue),
                "accrual_rows": pm.rows_accrual,
                "payment_rows": pm.rows_paid,
                "accrual_date_span": date_span(pm.filled_dates),
                "payment_date_span": date_span(pm.paid_dates),
                "data_quality": "partial_month" if month == "2026-05" else "ok",
            }
        )

    payroll_fields = list(payroll_rows[0].keys())
    write_csv(out_dir / "payroll_monthly.csv", payroll_rows, payroll_fields)

    courier_rows: list[dict[str, Any]] = []
    for month in COMPARISON_MONTHS:
        cm = couriers.get(month, CourierMonth())
        revenue_total = sales.get(month, {}).get("revenue", 0.0)
        orders_total = sales.get(month, {}).get("orders", 0.0)
        delivery_orders_iiko = delivery.get(month, {}).get("delivery_orders_iiko", 0.0)
        delivery_revenue_iiko = delivery.get(month, {}).get("delivery_revenue_iiko", 0.0)
        unique_orders = len(cm.unique_delivery_orders)
        courier_rows.append(
            {
                "month": month,
                "courier_payout_total": "",
                "courier_payout_source_status": "not_found_in_sheet",
                "delivery_rows_sheet": cm.delivery_rows,
                "unique_delivery_orders_sheet": unique_orders,
                "completed_delivery_rows_sheet": cm.completed_delivery_rows,
                "delivery_duration_hours_sum_sheet": round(cm.delivery_duration_hours_sum, 2),
                "shift_rows_sheet": cm.shift_rows,
                "completed_shift_rows_sheet": cm.completed_shift_rows,
                "shift_hours_sum_sheet": round(cm.shift_hours_sum, 2),
                "help_shift_rows_sheet": cm.help_shift_rows,
                "schedule_went_out_count_sheet": cm.schedule_status["Вышел"] + cm.schedule_status["Вышел2"] + cm.schedule_status["Помог"],
                "schedule_no_show_count_sheet": cm.schedule_status["Не вышел"] + cm.schedule_status["Не вышел2"],
                "cost_per_delivery_order": "",
                "courier_cost_pct_of_delivery_revenue": "",
                "iiko_total_revenue": round_money(revenue_total),
                "iiko_total_orders": round(orders_total, 2),
                "iiko_delivery_revenue": round_money(delivery_revenue_iiko),
                "iiko_delivery_orders": round(delivery_orders_iiko, 2),
                "sheet_vs_iiko_delivery_orders_delta": round(unique_orders - delivery_orders_iiko, 2) if delivery_orders_iiko else "",
                "delivery_date_span_sheet": date_span(cm.dates_delivery),
                "shift_date_span_sheet": date_span(cm.dates_shift),
                "data_quality": "no_courier_rows_for_month" if cm.delivery_rows == 0 and cm.shift_rows == 0 else "payout_rules_missing",
            }
        )

    courier_fields = list(courier_rows[0].keys())
    write_csv(out_dir / "couriers_monthly.csv", courier_rows, courier_fields)

    summary_rows: list[dict[str, Any]] = []
    for p_row, c_row in zip(payroll_rows, courier_rows, strict=True):
        summary_rows.append(
            {
                "month": p_row["month"],
                "iiko_revenue": p_row["iiko_revenue"],
                "payroll_accrual_total": p_row["payroll_accrual_total"],
                "payroll_accrual_pct_of_revenue": p_row["payroll_accrual_pct_of_revenue"],
                "payroll_cash_paid_total": p_row["payroll_cash_paid_total"],
                "central_admin_accrual": p_row["central_admin_accrual"],
                "shift_admin_accrual": p_row["shift_admin_accrual"],
                "kitchen_accrual": p_row["kitchen_accrual"],
                "courier_payout_total": c_row["courier_payout_total"],
                "courier_payout_source_status": c_row["courier_payout_source_status"],
                "unique_delivery_orders_sheet": c_row["unique_delivery_orders_sheet"],
                "iiko_delivery_orders": c_row["iiko_delivery_orders"],
                "iiko_delivery_revenue": c_row["iiko_delivery_revenue"],
                "labor_data_quality": "; ".join([p_row["data_quality"], c_row["data_quality"]]),
            }
        )
    write_csv(out_dir / "labor_cost_summary.csv", summary_rows, list(summary_rows[0].keys()))

    risks = [
        {
            "risk_id": "labor_001",
            "area": "payroll",
            "source": "Расчет зарплат NEW / Выгрузка",
            "period": "2025-11..2026-05",
            "risk": "Выгрузка хранит начисления построчно; удержания, депозиты и НДФЛ идут положительными строками и не должны суммироваться в ФОТ без классификации.",
            "action": "Использовать payroll_accrual_total как ФОТ по типам начислений; сверить трактовку накопительного фонда, депозитов и НДФЛ с владельцем.",
            "severity": "high",
        },
        {
            "risk_id": "labor_002",
            "area": "payroll",
            "source": "Расчет зарплат NEW / Выплаты",
            "period": "2025-11..2026-05",
            "risk": "Кассовые выплаты не равны расходу месяца начисления: есть авансы и выплаты накоплений.",
            "action": "В управленческом P&L брать начисления; выплаты использовать как cash-flow сверку.",
            "severity": "medium",
        },
        {
            "risk_id": "labor_003",
            "area": "couriers",
            "source": "График курьеров",
            "period": "2026-02..2026-05",
            "risk": "В книге найдены доставки, выходы и статистика, но не найдены суммы выплат или ставки начисления курьерской оплаты.",
            "action": "С владельцем подтвердить, где фиксируется ставка/выплата курьера: отдельная таблица, ДДС, касса или ручной расчет.",
            "severity": "high",
        },
        {
            "risk_id": "labor_004",
            "area": "couriers",
            "source": "График курьеров / Доставки и Выходы",
            "period": "2026-05",
            "risk": "В экспортированной курьерской книге данные по доставкам и выходам заканчиваются 2026-04-21; майских строк нет.",
            "action": "Проверить загрузку майских доставок/выходов в Google Sheets.",
            "severity": "high",
        },
        {
            "risk_id": "labor_005",
            "area": "iiko",
            "source": "iiko /reports/delivery/couriers",
            "period": "2025-11..2026-05",
            "risk": "Курьерский endpoint содержит строки AVERAGE/MAXIMUM/TARGET; их нельзя суммировать как факт заказов или выплат.",
            "action": "Использовать iiko courier endpoint только как вторичный ориентир по метрикам, не для ФОТ.",
            "severity": "medium",
        },
        {
            "risk_id": "labor_006",
            "area": "iiko",
            "source": "iiko employees attendance/schedule/salary",
            "period": "2026-02..2026-05",
            "risk": "iiko employee endpoints не дали пригодный факт часов, графика и выплат.",
            "action": "Считать Google Sheets источником факта ФОТ до появления надежной интеграции.",
            "severity": "medium",
        },
    ]
    write_csv(out_dir / "labor_quality_risks.csv", risks, list(risks[0].keys()))

    report = render_report(payroll_rows, courier_rows, summary_rows, risks, payroll_structure, courier_structure)
    (out_dir / "labor_report.md").write_text(report, encoding="utf-8")


def money(value: Any) -> str:
    if value == "" or value is None:
        return "н/д"
    return f"{float(value):,.0f}".replace(",", " ")


def percent_text(value: Any) -> str:
    if value == "" or value is None:
        return "н/д"
    return f"{float(value) * 100:.1f}%"


def render_report(
    payroll_rows: list[dict[str, Any]],
    courier_rows: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
    risks: list[dict[str, str]],
    payroll_structure: dict[str, Any],
    courier_structure: dict[str, Any],
) -> str:
    focus_payroll = [r for r in payroll_rows if r["month"] in FOCUS_MONTHS]
    focus_couriers = [r for r in courier_rows if r["month"] in FOCUS_MONTHS]

    payroll_table = [
        "| Месяц | Выручка iiko | ФОТ начисл. | ФОТ % | Центр. адм. | Сменные адм. | Кухня | Выплачено cash |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for r in focus_payroll:
        payroll_table.append(
            f"| {r['month']} | {money(r['iiko_revenue'])} | {money(r['payroll_accrual_total'])} | "
            f"{percent_text(r['payroll_accrual_pct_of_revenue'])} | {money(r['central_admin_accrual'])} | "
            f"{money(r['shift_admin_accrual'])} | {money(r['kitchen_accrual'])} | {money(r['payroll_cash_paid_total'])} |"
        )

    courier_table = [
        "| Месяц | Выплаты курьерам | Доставки Sheet, уник. | Доставки iiko | Доставка выручка iiko | Часы выходов | Стоимость на заказ |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for r in focus_couriers:
        courier_table.append(
            f"| {r['month']} | н/д | {r['unique_delivery_orders_sheet']} | {r['iiko_delivery_orders']} | "
            f"{money(r['iiko_delivery_revenue'])} | {r['shift_hours_sum_sheet']} | н/д |"
        )

    filled_months = [r["month"] for r in payroll_rows if r["accrual_rows"] > 0]
    courier_months = [r["month"] for r in courier_rows if r["delivery_rows_sheet"] > 0 or r["shift_rows_sheet"] > 0]
    sheet_list = ", ".join(payroll_structure["sheets"])
    courier_sheet_list = ", ".join(courier_structure["sheets"])

    lines = [
        "# ФОТ, курьеры и персонал",
        "",
        "Сформировано по локальным iiko-файлам и read-only экспортам Google Sheets. Персональные строки сотрудников и курьеров в отчет и CSV не вынесены.",
        "",
        "## Структура источников",
        "",
        f"- Факт / Источник: `Расчет зарплат NEW` / Период: 2025-11 — 2026-05 / Вывод: в книге есть листы {sheet_list}. Фактические начисления живут в `Выгрузка`, кассовые выплаты — в `Выплаты`, правила ставок и штат — в `Исходные данные`, `Штат`, `Категории и надбавки`, текущий расчет смен — в `Калькулятор`, смены и выручка — в `Смены и выручка`. / Действие: для P&L использовать `Выгрузка`, для cash-flow сверки использовать `Выплаты`.",
        f"- Факт / Источник: `График курьеров` / Период: 2026-02 — 2026-04 / Вывод: в книге есть листы {courier_sheet_list}. Доставки живут в `Доставки`, фактические выходы и часы — в `Выходы`, визуальный график — в `График`, агрегатная статистика — в `Статистика`. Сумм выплат и ставок начисления в книге не найдено. / Действие: подтвердить отдельный источник правил и факта курьерских выплат.",
        "",
        "## ФОТ",
        "",
        f"- Факт / Источник: `Расчет зарплат NEW / Выгрузка` / Период: {', '.join(filled_months)} / Вывод: месяцы 2025-11 — 2026-05 заполнены, май частичный по датам начисления. / Действие: май 2026 использовать как предварительный до закрытия месяца.",
        "",
        *payroll_table,
        "",
        "- Факт / Источник: `Расчет зарплат NEW / Выгрузка` / Период: 2026-02 — 2026-05 / Вывод: ФОТ рассчитан по начислительным типам `Оклад`, `Процент`, `Премия`, `Накопительный фонд`, `Больничные и отпуска и пособия`; удержания, НДФЛ и депозиты вынесены отдельно и не добавлены в ФОТ. / Действие: вручную подтвердить трактовку накопительного фонда как части ФОТ.",
        "- Факт / Источник: `Расчет зарплат NEW / Выплаты` / Период: 2026-02 — 2026-05 / Вывод: выплаты являются кассовым потоком и могут включать авансы/накопления, поэтому не равны расходу месяца начисления. / Действие: не смешивать `payroll_accrual_total` и `payroll_cash_paid_total` в одном P&L показателе.",
        "",
        "## Курьеры",
        "",
        f"- Факт / Источник: `График курьеров / Доставки, Выходы` / Период: {', '.join(courier_months)} / Вывод: заполнены фактические курьерские доставки/выходы до 2026-04-21; майских строк в экспортированной книге нет. / Действие: проверить загрузку майских курьерских данных.",
        "",
        *courier_table,
        "",
        "- Факт / Источник: `График курьеров` / Период: 2026-02 — 2026-05 / Вывод: курьерские выплаты, стоимость доставки на заказ и % от выручки доставки надежно посчитать нельзя, потому что суммы выплат и правила начисления не найдены. / Действие: сверить с владельцем, где хранится ставка/факт оплаты курьеров.",
        "- Факт / Источник: iiko `orders_delivery` / Период: 2025-11 — 2026-05 / Вывод: iiko дает выручку и заказы доставки по `Delivery.ServiceType=COURIER`, но не дает факт выплат курьерам. / Действие: использовать iiko только как знаменатель для % и сверку количества заказов.",
        "",
        "## Ненадежные данные",
        "",
    ]
    for risk in risks:
        lines.append(
            f"- Факт / Источник: {risk['source']} / Период: {risk['period']} / "
            f"Вывод: {risk['risk']} / Действие: {risk['action']}"
        )

    lines.extend(
        [
            "",
            "## Что сверить вручную",
            "",
            "- Факт / Источник: Google Sheets / Период: 2026-02 — 2026-05 / Вывод: курьерские выплаты отсутствуют в найденных листах. / Действие: запросить у владельца источник выплат или правило расчета: ставка за смену, ставка за заказ, доплаты, штрафы.",
            "- Факт / Источник: `Расчет зарплат NEW` / Период: 2026-02 — 2026-05 / Вывод: накопительный фонд, депозиты, НДФЛ и штрафы требуют управленческой трактовки. / Действие: закрепить, что входит в ФОТ, что является удержанием, а что cash-flow.",
            "- Факт / Источник: `График курьеров / Доставки` vs iiko / Период: 2026-02 — 2026-04 / Вывод: количество доставок в Sheet и iiko может расходиться из-за повторных строк на один заказ и статусов самовывоза/передачи курьера. / Действие: определить канонический счетчик доставок для стоимости на заказ.",
            "",
            "## Файлы",
            "",
            "- `research/processed/economic_block/payroll_monthly.csv`",
            "- `research/processed/economic_block/couriers_monthly.csv`",
            "- `research/processed/economic_block/labor_cost_summary.csv`",
            "- `research/processed/economic_block/labor_quality_risks.csv`",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--payroll-xlsx", required=True)
    parser.add_argument("--couriers-xlsx", required=True)
    parser.add_argument("--sales-daily", default="research/processed/iiko/sales/sales_daily.csv")
    parser.add_argument("--iiko-orders-raw-dir", default="research/raw/iiko/orders_delivery")
    parser.add_argument("--out-dir", default="research/processed/economic_block")
    args = parser.parse_args()
    build_outputs(args)


if __name__ == "__main__":
    main()
