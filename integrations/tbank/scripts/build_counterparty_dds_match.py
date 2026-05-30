#!/usr/bin/env python3
"""Match T-Bank counterparties to the classic DDS workbook.

Raw bank rows and full requisites stay in research/private/. Processed outputs keep
aggregates useful for owner review without bank accounts or full payment text.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import re
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[3]
TBANK_RAW = PROJECT_ROOT / "research/private/tbank/statement_2026-01-01_2026-03-31_p01.json"
DDS_WORKBOOK = PROJECT_ROOT / "research/private/google_drive/dds_2026_classic.xlsx"
PRIVATE_OUT = PROJECT_ROOT / "research/private/tbank/counterparties_2026-01_2026-03_dds_match_private.csv"
PROCESSED_OUT = PROJECT_ROOT / "research/processed/tbank/counterparties_2026-01_2026-03_dds_match.csv"
REPORT_OUT = PROJECT_ROOT / "research/processed/tbank/counterparties_2026-01_2026-03_dds_match_report.md"

START_DATE = dt.date(2026, 1, 1)
END_DATE = dt.date(2026, 3, 31)
MAX_DDS_DATE_LAG_DAYS = 5

Money = Decimal


LEGAL_OVERRIDES = [
    ("тора", "Амай", "Оплата поставщикам", "owner_confirmed"),
    ("альянс юг", "Альянс Юг", "Оплата поставщикам", "owner_confirmed"),
    ("мяснофф-дон", "Мяснов", "Оплата поставщикам", "owner_confirmed"),
    ("мистерия", "Мистерия", "Оплата поставщикам", "owner_confirmed"),
    ("метро кэш энд керри", "Метро", "Оплата поставщикам", "owner_confirmed"),
    ("о. о", "Синапс", "Контекстная реклама", "owner_confirmed"),
    ("айко", "ООО Айко", "Оплаты систем автоматизации", "owner_confirmed"),
    ("ревви", "Ревви", "Оплаты систем автоматизации", "owner_confirmed"),
    ("доксинбокс", "Докс ин бокс", "Оплаты систем автоматизации", "owner_confirmed"),
    ("синапсис", "ИП Трубина И.О", "SEO-оптимизация", "owner_confirmed_rename_to_Синопсис"),
    ("экоцентр", "Экоцентр", "Аренда торговых точек", "owner_confirmed"),
    ("суши принт", "Суши Принт", "Оплата поставщикам", "owner_confirmed"),
    ("сити", "Частные подрядчики", "Баннерная реклама", "owner_confirmed_legacy_supplier"),
]

CARD_MERCHANT_RULES = [
    (("ozon",), "Озон", "", "merchant_rule"),
    (("mango-office", "mango"), "Манго Телеком", "Телекоммуникации", "merchant_rule"),
    (("avito",), "Авито", "Поиск и найм персонала", "merchant_rule"),
    (("ihc",), "ihs", "Телекоммуникации", "merchant_rule"),
    (("upakcentr",), "Упак", "Оплата поставщикам", "merchant_rule"),
    (
        (
            "magnit",
            "pyaterochka",
            "magazin",
            "ovoshhi",
            "gorod sad",
            "sukhofrukty",
            "paprika",
            "pobeda",
            "optovyj",
        ),
        "Розничные поставщики",
        "Оплата поставщикам",
        "merchant_rule",
    ),
]


def money(value: Any) -> Money:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, str):
        value = value.replace(" ", "").replace(",", ".")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def fmt_money(value: Money) -> str:
    return str(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def clean_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def normalize(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip().casefold()


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def sha(value: str, length: int = 12) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:length]


def inn_mask(value: str) -> str:
    digits = clean_digits(value)
    if len(digits) < 4:
        return ""
    return f"{'*' * (len(digits) - 4)}{digits[-4:]}"


def inn_kind(value: str) -> str:
    digits = clean_digits(value)
    if len(digits) == 10:
        return "legal_10"
    if len(digits) == 12:
        return "person_or_ip_12"
    if digits:
        return f"other_{len(digits)}"
    return "missing"


def operation_date(row: dict[str, Any]) -> dt.date:
    raw = str(row.get("operationDate") or row.get("docDate") or row.get("chargeDate") or "")
    match = re.search(r"\d{4}-\d{2}-\d{2}", raw)
    if not match:
        return START_DATE
    return dt.date.fromisoformat(match.group(0))


def operation_amount(row: dict[str, Any]) -> Money:
    return abs(money(row.get("rubleAmount") or row.get("accountAmount") or row.get("operationAmount"))).quantize(
        Decimal("0.01")
    )


def operation_direction(row: dict[str, Any]) -> str:
    return "credit" if str(row.get("typeOfOperation") or "").casefold() == "credit" else "debit"


def dds_sign_for_direction(direction: str) -> str:
    return "Поступление" if direction == "credit" else "Выбытие"


def counterparty_block(row: dict[str, Any]) -> dict[str, Any]:
    direction = operation_direction(row)
    block = row.get("payer") if direction == "credit" else row.get("receiver")
    if isinstance(block, dict):
        return block
    block = row.get("counterParty")
    return block if isinstance(block, dict) else {}


def block_value(block: dict[str, Any], names: tuple[str, ...]) -> str:
    for name in names:
        value = block.get(name)
        if value not in (None, "") and not isinstance(value, (dict, list)):
            return str(value)
    return ""


CARD_MERCHANT_RE = re.compile(r"номер\s+\S+\s+(.+?)(?:\.\s*Договор| Договор|$)", re.I)


def card_merchant(row: dict[str, Any]) -> str:
    if str(row.get("category") or "") != "cardOperation":
        return ""
    text = str(row.get("payPurpose") or row.get("description") or "")
    match = CARD_MERCHANT_RE.search(text)
    merchant = match.group(1) if match else text
    return re.sub(r"\s+", " ", merchant).strip()


def card_merchant_family(merchant: str) -> str:
    low = normalize(merchant)
    families = [
        ("MAGNIT", ("magnit",)),
        ("PYATEROCHKA", ("pyaterochka",)),
        ("OZON", ("ozon",)),
        ("MANGO-OFFICE", ("mango-office",)),
        ("AVITO", ("avito",)),
        ("UPAKCENTR", ("upakcentr",)),
        ("IHC", ("ihc",)),
    ]
    for family, needles in families:
        if any(needle in low for needle in needles):
            return family
    return merchant


def logical_counterparty(row: dict[str, Any]) -> dict[str, str]:
    block = counterparty_block(row)
    merchant = card_merchant(row)
    if merchant:
        name = card_merchant_family(merchant)
        source = "card_merchant"
    else:
        name = " ".join(block_value(block, ("name", "fullName", "shortName")).split())
        source = "payer_receiver"
    inn = clean_digits(block_value(block, ("inn", "INN", "taxId", "tin")))
    account = clean_digits(block_value(block, ("acct", "account", "accountNumber", "bankAccount")))
    return {
        "source": source,
        "name": name,
        "raw_bank_name": " ".join(block_value(block, ("name", "fullName", "shortName")).split()),
        "inn": inn,
        "account": account,
        "merchant": merchant,
        "merchant_family": card_merchant_family(merchant) if merchant else "",
    }


def load_tbank_operations(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows = payload.get("operations", []) if isinstance(payload, dict) else []
    result = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        day = operation_date(row)
        if START_DATE <= day <= END_DATE:
            result.append(row)
    return result


def load_dds_rows(path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["ДДС месяц"]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=4, max_row=4))]
    idx = {name: pos for pos, name in enumerate(header) if name}
    rows = []
    for row_num, raw in enumerate(worksheet.iter_rows(min_row=5, values_only=True), start=5):
        values = list(raw)
        raw_date = values[idx["Дата"]] if idx["Дата"] < len(values) else None
        if not hasattr(raw_date, "date"):
            continue
        day = raw_date.date()
        if not (START_DATE <= day <= END_DATE):
            continue
        amount_value = values[idx["Сумма"]] if idx["Сумма"] < len(values) else None
        if amount_value in (None, ""):
            continue
        signed = money(amount_value)
        rows.append(
            {
                "row_num": row_num,
                "date": day,
                "amount": abs(signed).quantize(Decimal("0.01")),
                "sign": "Поступление" if signed > 0 else "Выбытие",
                "wallet": str(values[idx["Кошелек"]] or "").strip(),
                "business_direction": str(values[idx["Направление бизнеса"]] or "").strip(),
                "counterparty": str(values[idx["Контрагент"]] or "").strip(),
                "payment_purpose": str(values[idx["Назначение платежа"]] or "").strip(),
                "article": str(values[idx["Статья"]] or "").strip(),
                "activity_type": str(values[idx["Вид д-ти"]] or "").strip(),
            }
        )
    return rows


def legal_override(name: str) -> tuple[str, str, str] | None:
    low = normalize(name)
    for needle, dds_counterparty, article, status in LEGAL_OVERRIDES:
        if needle in low:
            return dds_counterparty, article, status
    return None


def merchant_override(merchant: str) -> tuple[str, str, str] | None:
    low = normalize(merchant)
    for needles, dds_counterparty, article, status in CARD_MERCHANT_RULES:
        if any(needle in low for needle in needles):
            return dds_counterparty, article, status
    return None


def name_score(row: dict[str, Any], candidate: dict[str, Any]) -> int:
    cp = logical_counterparty(row)
    dds_cp = normalize(candidate["counterparty"])
    dds_article = normalize(candidate["article"])
    score = 0

    override = legal_override(cp["raw_bank_name"] or cp["name"])
    if override and normalize(override[0]) == dds_cp:
        score += 50
    merchant = cp["merchant"] or cp["name"]
    merchant_rule = merchant_override(merchant)
    if merchant_rule and normalize(merchant_rule[0]) == dds_cp:
        score += 35

    words = [word for word in re.split(r"[^а-яa-z0-9]+", normalize(cp["name"])) if len(word) >= 4]
    for word in words:
        if word in dds_cp:
            score += 8

    category = str(row.get("category") or "")
    if category in {"tax", "budget"} and "налог" in dds_cp:
        score += 40
    if category == "fee" and "банк" in dds_cp:
        score += 30
    if category == "incomeLoan" and "получение овердрафта" in dds_article:
        score += 60
    if category == "creditPaymentOuter" and "погашение овердрафта" in dds_article:
        score += 60
    if category == "refundIn" and "возврат" in dds_article:
        score += 30
    if category == "selfTransferOuter" and "перевод между счетами" in dds_article:
        score += 45
    return score


def match_operation(row: dict[str, Any], dds_by_amount: dict[tuple[Money, str], list[dict[str, Any]]]) -> dict[str, Any]:
    day = operation_date(row)
    amount_value = operation_amount(row)
    sign = dds_sign_for_direction(operation_direction(row))
    category = str(row.get("category") or "")
    candidates = [
        candidate
        for candidate in dds_by_amount.get((amount_value, sign), [])
        if abs((candidate["date"] - day).days) <= MAX_DDS_DATE_LAG_DAYS
    ]
    if category == "incomeLoan":
        candidates = [candidate for candidate in candidates if "овердрафт" in normalize(candidate["article"])]
    elif category == "creditPaymentOuter":
        candidates = [candidate for candidate in candidates if "овердрафт" in normalize(candidate["article"])]
    elif category == "refundIn":
        candidates = [candidate for candidate in candidates if "возврат" in normalize(candidate["article"])]
    elif category in {"tax", "budget"}:
        candidates = [candidate for candidate in candidates if "налог" in normalize(candidate["article"])]
    if candidates:
        candidates.sort(
            key=lambda candidate: (
                -name_score(row, candidate),
                abs((candidate["date"] - day).days),
                0 if candidate["counterparty"] else 1,
                candidate["row_num"],
            )
        )
        chosen = candidates[0]
        return {
            "status": "matched_amount_date",
            "dds_counterparty": chosen["counterparty"],
            "dds_article": chosen["article"],
            "dds_date": chosen["date"].isoformat(),
            "dds_row_num": chosen["row_num"],
            "date_lag_days": (chosen["date"] - day).days,
        }

    cp = logical_counterparty(row)
    override = legal_override(cp["raw_bank_name"] or cp["name"])
    if override:
        return {
            "status": override[2],
            "dds_counterparty": override[0],
            "dds_article": override[1],
            "dds_date": "",
            "dds_row_num": "",
            "date_lag_days": "",
        }
    merchant_rule = merchant_override(cp["merchant"] or cp["name"])
    if merchant_rule:
        return {
            "status": merchant_rule[2],
            "dds_counterparty": merchant_rule[0],
            "dds_article": merchant_rule[1],
            "dds_date": "",
            "dds_row_num": "",
            "date_lag_days": "",
        }
    if category == "incomeLoan":
        return {
            "status": "category_rule",
            "dds_counterparty": "Банки",
            "dds_article": "Получение овердрафта",
            "dds_date": "",
            "dds_row_num": "",
            "date_lag_days": "",
        }
    if category == "creditPaymentOuter":
        return {
            "status": "category_rule",
            "dds_counterparty": "Банки",
            "dds_article": "Погашение овердрафта",
            "dds_date": "",
            "dds_row_num": "",
            "date_lag_days": "",
        }
    if category in {"tax", "budget"}:
        return {
            "status": "category_rule",
            "dds_counterparty": "Налоговая",
            "dds_article": "Налоги",
            "dds_date": "",
            "dds_row_num": "",
            "date_lag_days": "",
        }
    return {
        "status": "unmatched",
        "dds_counterparty": "",
        "dds_article": "",
        "dds_date": "",
        "dds_row_num": "",
        "date_lag_days": "",
    }


def sample_purpose(row: dict[str, Any]) -> str:
    return re.sub(r"\s+", " ", str(row.get("payPurpose") or row.get("description") or "")).strip()[:260]


def build_rows(operations: list[dict[str, Any]], dds_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    dds_by_amount: dict[tuple[Money, str], list[dict[str, Any]]] = defaultdict(list)
    for row in dds_rows:
        dds_by_amount[(row["amount"], row["sign"])].append(row)

    operation_matches = []
    for row in operations:
        cp = logical_counterparty(row)
        match = match_operation(row, dds_by_amount)
        operation_matches.append(
            {
                "row": row,
                "cp": cp,
                "match": match,
            }
        )

    groups: dict[tuple[str, str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for item in operation_matches:
        row = item["row"]
        cp = item["cp"]
        key = (
            operation_direction(row),
            str(row.get("category") or ""),
            cp["source"],
            cp["inn"],
            cp["name"],
        )
        groups[key].append(item)

    private_rows = []
    processed_rows = []
    for key, items in sorted(groups.items(), key=lambda kv: sum(operation_amount(item["row"]) for item in kv[1]), reverse=True):
        direction, category, source, inn, name = key
        total = sum((operation_amount(item["row"]) for item in items), Decimal("0"))
        matched_items = [item for item in items if item["match"]["status"] != "unmatched"]
        matched_total = sum((operation_amount(item["row"]) for item in matched_items), Decimal("0"))
        dates = sorted(operation_date(item["row"]).isoformat() for item in items)
        dds_counterparties = Counter(item["match"]["dds_counterparty"] or "" for item in matched_items)
        dds_articles = Counter(item["match"]["dds_article"] or "" for item in matched_items)
        statuses = Counter(item["match"]["status"] for item in items)
        bank_names = Counter(item["cp"]["raw_bank_name"] for item in items if item["cp"]["raw_bank_name"])
        merchants = Counter(item["cp"]["merchant_family"] or item["cp"]["merchant"] for item in items if item["cp"]["merchant"])
        counterparty_id = "TBDDS_" + sha("|".join(key))
        primary_dds_counterparty = dds_counterparties.most_common(1)[0][0] if dds_counterparties else ""
        primary_dds_article = dds_articles.most_common(1)[0][0] if dds_articles else ""
        private_row = {
            "counterparty_id": counterparty_id,
            "period": "2026-01..2026-03",
            "direction": direction,
            "tbank_category": category,
            "counterparty_source": source,
            "logical_counterparty_name": name,
            "bank_counterparty_names": "; ".join(name for name, _count in bank_names.most_common(6)),
            "card_merchants": "; ".join(name for name, _count in merchants.most_common(10)),
            "counterparty_inn": inn,
            "counterparty_inn_kind": inn_kind(inn),
            "operation_count": len(items),
            "amount_total_abs": fmt_money(total),
            "matched_operation_count": len(matched_items),
            "matched_amount_abs": fmt_money(matched_total),
            "unmatched_operation_count": len(items) - len(matched_items),
            "unmatched_amount_abs": fmt_money(total - matched_total),
            "first_date": dates[0] if dates else "",
            "last_date": dates[-1] if dates else "",
            "primary_dds_counterparty": primary_dds_counterparty,
            "primary_dds_article": primary_dds_article,
            "dds_counterparties": "; ".join(f"{name} ({count})" for name, count in dds_counterparties.most_common() if name),
            "dds_articles": "; ".join(f"{name} ({count})" for name, count in dds_articles.most_common() if name),
            "match_statuses": "; ".join(f"{status} ({count})" for status, count in statuses.most_common()),
            "sample_purpose_private": sample_purpose(items[0]["row"]),
        }
        private_rows.append(private_row)
        processed_rows.append(
            {
                "counterparty_id": counterparty_id,
                "period": private_row["period"],
                "direction": direction,
                "tbank_category": category,
                "counterparty_source": source,
                "logical_counterparty_name": name,
                "inn_kind": inn_kind(inn),
                "inn_mask": inn_mask(inn),
                "operation_count": len(items),
                "amount_total_abs": fmt_money(total),
                "matched_operation_count": len(matched_items),
                "matched_amount_abs": fmt_money(matched_total),
                "unmatched_operation_count": len(items) - len(matched_items),
                "unmatched_amount_abs": fmt_money(total - matched_total),
                "first_date": private_row["first_date"],
                "last_date": private_row["last_date"],
                "primary_dds_counterparty": primary_dds_counterparty,
                "primary_dds_article": primary_dds_article,
                "dds_counterparties": private_row["dds_counterparties"],
                "dds_articles": private_row["dds_articles"],
                "match_statuses": private_row["match_statuses"],
            }
        )

    detail_rows = []
    for item in operation_matches:
        row = item["row"]
        cp = item["cp"]
        match = item["match"]
        detail_rows.append(
            {
                "operation_id": row.get("operationId") or "",
                "operation_date": operation_date(row).isoformat(),
                "direction": operation_direction(row),
                "tbank_category": row.get("category") or "",
                "amount_abs": fmt_money(operation_amount(row)),
                "counterparty_source": cp["source"],
                "logical_counterparty_name": cp["name"],
                "bank_counterparty_name": cp["raw_bank_name"],
                "counterparty_inn": cp["inn"],
                "card_merchant": cp["merchant"],
                "match_status": match["status"],
                "dds_counterparty": match["dds_counterparty"],
                "dds_article": match["dds_article"],
                "dds_date": match["dds_date"],
                "dds_row_num": match["dds_row_num"],
                "date_lag_days": match["date_lag_days"],
                "sample_purpose_private": sample_purpose(row),
            }
        )
    return private_rows, processed_rows, detail_rows


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def write_report(processed_rows: list[dict[str, Any]], operations: list[dict[str, Any]], dds_rows: list[dict[str, Any]]) -> None:
    total_amount = sum((money(row["amount_total_abs"]) for row in processed_rows), Decimal("0"))
    matched_amount = sum((money(row["matched_amount_abs"]) for row in processed_rows), Decimal("0"))
    unmatched_amount = total_amount - matched_amount
    matched_ops = sum(int(row["matched_operation_count"]) for row in processed_rows)
    unmatched_ops = sum(int(row["unmatched_operation_count"]) for row in processed_rows)
    debit_rows = [row for row in processed_rows if row["direction"] == "debit"]
    top_debit = sorted(debit_rows, key=lambda row: money(row["amount_total_abs"]), reverse=True)[:25]
    unmatched_top = sorted(
        [row for row in processed_rows if money(row["unmatched_amount_abs"]) > 0],
        key=lambda row: money(row["unmatched_amount_abs"]),
        reverse=True,
    )[:15]

    lines = [
        "# T-Bank counterparties matched to DDS",
        "",
        f"Период T-Bank API: `{START_DATE.isoformat()}` - `{END_DATE.isoformat()}`.",
        f"Операций T-Bank: {len(operations)}.",
        f"Строк DDS в периоде: {len(dds_rows)}.",
        f"Уникальных логических контрагентов/категорий T-Bank: {len(processed_rows)}.",
        "",
        "## Покрытие",
        "",
        f"- Сопоставлено операций: {matched_ops}.",
        f"- Не сопоставлено операций: {unmatched_ops}.",
        f"- Сопоставлено по сумме: {fmt_money(matched_amount)} руб.",
        f"- Не сопоставлено по сумме: {fmt_money(unmatched_amount)} руб.",
        "",
        "## Топ Расходных Контрагентов",
        "",
        "| Контрагент | Категория T-Bank | Сумма | DDS контрагент | DDS статьи | Статусы |",
        "| --- | --- | ---: | --- | --- | --- |",
    ]
    for row in top_debit:
        lines.append(
            f"| `{row['logical_counterparty_name']}` | `{row['tbank_category']}` | {row['amount_total_abs']} | "
            f"{row['primary_dds_counterparty']} | {row['dds_articles']} | {row['match_statuses']} |"
        )
    lines.extend(
        [
            "",
            "## Требует Доработки",
            "",
            "| Контрагент | Категория T-Bank | Непокрытая сумма | Операций | Подсказка |",
            "| --- | --- | ---: | ---: | --- |",
        ]
    )
    for row in unmatched_top:
        hint = "бизнес-карта: в DDS часто заведено агрегатом" if row["counterparty_source"] == "card_merchant" else "нет точного совпадения сумма+дата в DDS"
        lines.append(
            f"| `{row['logical_counterparty_name']}` | `{row['tbank_category']}` | "
            f"{row['unmatched_amount_abs']} | {row['unmatched_operation_count']} | {hint} |"
        )
    lines.extend(
        [
            "",
            "## Файлы",
            "",
            f"- `{rel(PROCESSED_OUT)}`",
            f"- `{rel(PRIVATE_OUT)}`",
            f"- `research/private/tbank/counterparties_2026-01_2026-03_dds_operation_match_private.csv`",
            "",
        ]
    )
    REPORT_OUT.parent.mkdir(parents=True, exist_ok=True)
    REPORT_OUT.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    operations = load_tbank_operations(TBANK_RAW)
    dds_rows = load_dds_rows(DDS_WORKBOOK)
    private_rows, processed_rows, detail_rows = build_rows(operations, dds_rows)

    private_fields = [
        "counterparty_id",
        "period",
        "direction",
        "tbank_category",
        "counterparty_source",
        "logical_counterparty_name",
        "bank_counterparty_names",
        "card_merchants",
        "counterparty_inn",
        "counterparty_inn_kind",
        "operation_count",
        "amount_total_abs",
        "matched_operation_count",
        "matched_amount_abs",
        "unmatched_operation_count",
        "unmatched_amount_abs",
        "first_date",
        "last_date",
        "primary_dds_counterparty",
        "primary_dds_article",
        "dds_counterparties",
        "dds_articles",
        "match_statuses",
        "sample_purpose_private",
    ]
    processed_fields = [
        "counterparty_id",
        "period",
        "direction",
        "tbank_category",
        "counterparty_source",
        "logical_counterparty_name",
        "inn_kind",
        "inn_mask",
        "operation_count",
        "amount_total_abs",
        "matched_operation_count",
        "matched_amount_abs",
        "unmatched_operation_count",
        "unmatched_amount_abs",
        "first_date",
        "last_date",
        "primary_dds_counterparty",
        "primary_dds_article",
        "dds_counterparties",
        "dds_articles",
        "match_statuses",
    ]
    detail_fields = [
        "operation_id",
        "operation_date",
        "direction",
        "tbank_category",
        "amount_abs",
        "counterparty_source",
        "logical_counterparty_name",
        "bank_counterparty_name",
        "counterparty_inn",
        "card_merchant",
        "match_status",
        "dds_counterparty",
        "dds_article",
        "dds_date",
        "dds_row_num",
        "date_lag_days",
        "sample_purpose_private",
    ]

    write_csv(PRIVATE_OUT, private_rows, private_fields)
    write_csv(PROCESSED_OUT, processed_rows, processed_fields)
    write_csv(
        PROJECT_ROOT / "research/private/tbank/counterparties_2026-01_2026-03_dds_operation_match_private.csv",
        detail_rows,
        detail_fields,
    )
    write_report(processed_rows, operations, dds_rows)

    print(f"operations={len(operations)}")
    print(f"dds_rows={len(dds_rows)}")
    print(f"counterparty_groups={len(processed_rows)}")
    print(f"processed={rel(PROCESSED_OUT)}")
    print(f"private={rel(PRIVATE_OUT)}")
    print(f"report={rel(REPORT_OUT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
